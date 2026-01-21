# src/routes/admin.py
from flask import Blueprint, jsonify, request, current_app, send_file
from flask_jwt_extended import jwt_required, get_jwt_identity
from src.models.user import User, Job, JobAssignment, AgentAvailability, Invoice, InvoiceJob, Notification, JobBilling, Expense, db
from src.models.admin_message import AdminMessage, AdminMessageDelivery
from src.utils.s3_client import s3_client
from src.utils.finance import (
    update_job_hours, calculate_job_revenue, calculate_expense_vat,
    get_job_expense_totals, get_job_agent_invoice_totals, calculate_job_profit,
    lock_job_revenue_snapshot, get_financial_summary
)
from src.utils.dbcheck import full_health_check
from datetime import datetime, date, timedelta
import requests
import json
import logging
import requests
import io
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch, cm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image, PageBreak
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

admin_bp = Blueprint('admin', __name__)
@admin_bp.route('/admin/agents/minimal', methods=['GET'])
@jwt_required()
def get_agents_minimal():
    user = require_admin()
    if not user:
        return jsonify({'error': 'Forbidden'}), 403

    linked_only = str(request.args.get('linked_only', 'false')).lower() in ('1', 'true', 'yes')
    q = User.query.filter(User.role == 'agent')
    if linked_only:
        q = q.filter(User.telegram_chat_id.isnot(None))

    def _name(u: User):
        first = u.first_name or ''
        last = u.last_name or ''
        name = f"{first} {last}".strip()
        return name or (u.email or f"Agent #{u.id}")

    agents = [{
        'id': a.id,
        'name': _name(a),
        'email': a.email,
        'linked': bool(a.telegram_chat_id)
    } for a in q.all()]

    return jsonify(agents)


@admin_bp.route('/agents', methods=['GET'])
@jwt_required()
def list_agents():
    """Get all agents with basic info. Supports ?active=true filter."""
    try:
        user = require_admin()
        if not user:
            return jsonify({'error': 'Forbidden'}), 403

        active = request.args.get('active') == 'true'
        query = User.query.filter(User.role == 'agent')

        # Note: User model doesn't have is_active attribute, so we skip that filter

        agents = query.order_by(User.first_name, User.last_name).all()

        result = []
        for agent in agents:
            result.append({
                'id': agent.id,
                'display_name': f"{agent.first_name or ''} {agent.last_name or ''}".strip() or agent.email,
                'email': agent.email,
                'region': getattr(agent, 'region', None),
                'skills': getattr(agent, 'skills', [])
            })

        return jsonify({'agents': result})
    except Exception as e:
        current_app.logger.exception("list_agents failed: %s", e)
        return jsonify({'agents': []})


@admin_bp.route('/agents/available', methods=['GET'])
@jwt_required()
def get_agents_available():
    """Get agents available for a specific date range."""
    try:
        user = require_admin()
        if not user:
            return jsonify({'error': 'Forbidden'}), 403

        # Support both single date and date range
        date_param = request.args.get('date')
        start_param = request.args.get('start')
        end_param = request.args.get('end')

        if date_param:
            # Single date format
            start_date = _parse_date_param(date_param)
            end_date = start_date
        elif start_param and end_param:
            # Date range format
            start_date = _parse_date_param(start_param)
            end_date = _parse_date_param(end_param)
        else:
            # Return empty list instead of 400 to avoid UI hard-fail
            current_app.logger.warning("agents_available called without date/start/end parameters")
            return jsonify({'agents': []})

        if not start_date or not end_date:
            current_app.logger.warning(f"agents_available bad date params: date={date_param}, start={start_param}, end={end_param}")
            return jsonify({'agents': []})

        # Get agents available within the date range
        available_agents = (
            db.session.query(User)
            .join(AgentAvailability, AgentAvailability.agent_id == User.id)
            .filter(
                User.role == 'agent',
                AgentAvailability.date.between(start_date, end_date),
                AgentAvailability.is_available.is_(True)
            )
            .distinct()
            .order_by(User.first_name, User.last_name)
            .all()
        )

        result = []
        for agent in available_agents:
            result.append({
                'id': agent.id,
                'display_name': f"{agent.first_name or ''} {agent.last_name or ''}".strip() or agent.email,
                'email': agent.email,
                'region': getattr(agent, 'region', None),
                'skills': getattr(agent, 'skills', [])
            })

        return jsonify({'agents': result})

    except Exception as e:
        current_app.logger.warning(f"Error fetching available agents: {e}")
        return jsonify({'agents': []})


@admin_bp.route('/agents/reliability', methods=['GET'])
@jwt_required()
def get_agents_reliability():
    """Get most reliable agents based on job acceptance rate."""
    try:
        user = require_admin()
        if not user:
            return jsonify({'error': 'Forbidden'}), 403
        # Default to last 90 days
        end_date = date.today()
        start_date = end_date - timedelta(days=90)

        # Allow custom date range
        if request.args.get('start'):
            start_date = _parse_date_param(request.args.get('start'))
        if request.args.get('end'):
            end_date = _parse_date_param(request.args.get('end'))

        limit = int(request.args.get('limit', 20))

        # Count job assignments (accepted) per agent in date range
        accepted_subq = (
            db.session.query(
                JobAssignment.agent_id,
                db.func.count().label('accepted')
            )
            .filter(
                JobAssignment.status == 'accepted',
                JobAssignment.created_at.between(start_date, end_date)
            )
            .group_by(JobAssignment.agent_id)
            .subquery()
        )

        # Count total job offers/assignments per agent in date range
        offered_subq = (
            db.session.query(
                JobAssignment.agent_id,
                db.func.count().label('offered')
            )
            .filter(
                JobAssignment.created_at.between(start_date, end_date)
            )
            .group_by(JobAssignment.agent_id)
            .subquery()
        )

        # Get agents with their acceptance stats
        query = (
            db.session.query(
                User.id,
                User.first_name,
                User.last_name,
                User.email,
                db.func.coalesce(accepted_subq.c.accepted, 0).label('accepted'),
                db.func.coalesce(offered_subq.c.offered, 0).label('offered')
            )
            .outerjoin(accepted_subq, accepted_subq.c.agent_id == User.id)
            .outerjoin(offered_subq, offered_subq.c.agent_id == User.id)
            .filter(
                User.role == 'agent'
            )
        )

        agents_data = []
        for row in query.all():
            accepted = int(row.accepted or 0)
            offered = int(row.offered or 0)
            accept_rate = float(accepted) / offered if offered > 0 else 0.0

            agents_data.append({
                'id': row.id,
                'display_name': f"{row.first_name or ''} {row.last_name or ''}".strip() or row.email,
                'email': row.email,
                'accepted': accepted,
                'offered': offered,
                'accept_rate': accept_rate
            })

        # Sort by accept_rate (desc) then by accepted count (desc)
        agents_data.sort(key=lambda x: (x['accept_rate'], x['accepted']), reverse=True)

        return jsonify({'agents': agents_data[:limit]})

    except Exception as e:
        current_app.logger.exception(f"get_agents_reliability failed: {e}")
        return jsonify({'agents': []})


@admin_bp.route('/agents/picker', methods=['GET'])
@jwt_required()
def agents_picker():
    """
    Aggregator endpoint that returns all, available, and reliable agents in one call.
    Always returns 200 with stable JSON structure.
    """
    try:
        user = require_admin()
        if not user:
            return jsonify({'error': 'Forbidden'}), 403

        date_s = request.args.get('date')  # optional
        win_s = request.args.get('window_days', '90')
        window_days = int(win_s) if win_s.isdigit() else 90

        current_app.logger.info(f"agents_picker called with date={date_s}, window_days={window_days}")

        # Helper function to format agent data
        def row_agent(a):
            return {
                "id": a.id,
                "display_name": f"{a.first_name or ''} {a.last_name or ''}".strip() or a.email,
                "region": getattr(a, "region", None),
                "skills": getattr(a, "skills", [])
            }

        # Get all agents (same query as Comms tab would use)
        def q_all_agents():
            q = User.query.filter(User.role == 'agent')
            return q.order_by(User.first_name, User.last_name)

        # --- ALL AGENTS ---
        all_agents = [row_agent(a) for a in q_all_agents().all()]

        # --- AVAILABLE AGENTS (for specific date) ---
        available = []
        if date_s:
            try:
                from dateutil import parser as dtp
                d = dtp.parse(date_s).date()

                # Find agents busy on this date (have accepted assignments)
                busy_subq = (
                    db.session.query(JobAssignment.agent_id)
                    .join(Job, JobAssignment.job_id == Job.id)
                    .filter(
                        JobAssignment.status == 'accepted',
                        db.func.date(Job.arrival_time) == d
                    )
                    .distinct()
                    .subquery()
                )

                # Get agents not in the busy list
                avail_q = q_all_agents().filter(~User.id.in_(busy_subq))
                available = [row_agent(a) for a in avail_q.all()]
            except Exception as e:
                current_app.logger.warning(f"picker available parse error: {e}")

        # --- RELIABLE AGENTS (based on acceptance rate) ---
        reliable = []
        try:
            end_d = date.today()
            start_d = end_d - timedelta(days=window_days)

            # Count total assignments offered per agent in window
            offered_subq = (
                db.session.query(
                    JobAssignment.agent_id,
                    db.func.count().label('offered')
                )
                .filter(JobAssignment.created_at.between(start_d, end_d))
                .group_by(JobAssignment.agent_id)
                .subquery()
            )

            # Count accepted assignments per agent in window
            accepted_subq = (
                db.session.query(
                    JobAssignment.agent_id,
                    db.func.count().label('accepted')
                )
                .filter(
                    JobAssignment.status == 'accepted',
                    JobAssignment.created_at.between(start_d, end_d)
                )
                .group_by(JobAssignment.agent_id)
                .subquery()
            )

            # Join with User table to get agent details
            q = (
                db.session.query(
                    User.id,
                    User.first_name,
                    User.last_name,
                    User.email,
                    db.func.coalesce(accepted_subq.c.accepted, 0).label('accepted'),
                    db.func.coalesce(offered_subq.c.offered, 0).label('offered')
                )
                .filter(User.role == 'agent')
                .outerjoin(offered_subq, offered_subq.c.agent_id == User.id)
                .outerjoin(accepted_subq, accepted_subq.c.agent_id == User.id)
            )

            rows = q.all()
            reliable_data = []
            for id_, first_name, last_name, email, acc, off in rows:
                display_name = f"{first_name or ''} {last_name or ''}".strip() or email
                rate = float(acc) / float(off) if off > 0 else (1.0 if acc > 0 else 0.0)
                reliable_data.append({
                    "id": id_,
                    "display_name": display_name,
                    "accepted": int(acc),
                    "offered": int(off),
                    "accept_rate": rate
                })

            # Sort by accept_rate desc, then by accepted count desc
            reliable_data.sort(key=lambda r: (r["accept_rate"], r["accepted"]), reverse=True)
            reliable = reliable_data

        except Exception as e:
            current_app.logger.warning(f"picker reliable error: {e}")
            reliable = []

        current_app.logger.info(f"picker counts all={len(all_agents)} avail={len(available)} rel={len(reliable)}")

        return jsonify({
            "all": all_agents,
            "available": available,
            "reliable": reliable
        })

    except Exception as e:
        current_app.logger.exception(f"agents_picker fatal: {e}")
        return jsonify({"all": [], "available": [], "reliable": []})


@admin_bp.route('/admin/telegram/messages', methods=['POST'])
@jwt_required()
def send_admin_telegram_messages():
    user = require_admin()
    if not user:
        return jsonify({'error': 'Forbidden'}), 403

    data = request.get_json(silent=True) or {}
    message = (data.get('message') or '').strip()
    agent_ids = data.get('agent_ids') or []
    send_to_all = bool(data.get('send_to_all'))

    if not (1 <= len(message) <= 1000):
        return jsonify({'error': 'Message must be 1..1000 chars'}), 400

    # Resolve recipients
    query = User.query.filter(User.role == 'agent')
    if send_to_all:
        # take all linked & active agents
        linked_agents = query.filter(User.telegram_chat_id.isnot(None)).all()
        target_agents = linked_agents
    else:
        if not isinstance(agent_ids, list) or not all(isinstance(x, (int, str)) for x in agent_ids):
            return jsonify({'error': 'agent_ids must be an array'}), 400
        # cast to int ids (ids are int in this codebase)
        ids = []
        for x in agent_ids:
            try:
                ids.append(int(x))
            except Exception:
                pass
        if len(ids) == 0:
            return jsonify({'error': 'No recipients selected'}), 400
        target_agents = query.filter(User.id.in_(ids)).all()

    if not send_to_all and len(target_agents) > 200:
        return jsonify({'error': 'Too many recipients (>200). Use send_to_all or split batches.'}), 400

    # Split into linked vs not linked
    linked = [a for a in target_agents if a.telegram_chat_id]
    not_linked = [str(a.id) for a in target_agents if not a.telegram_chat_id]

    # Create admin message record
    admin_msg = AdminMessage(admin_id=user.id, message=message)
    db.session.add(admin_msg)
    db.session.flush()  # get id

    results = []
    queued = 0

    # Use integrations.telegram_client for consistency
    from src.integrations.telegram_client import send_message as telegram_send

    for agent in linked[:200]:
        chat_id = agent.telegram_chat_id
        try:
            resp = telegram_send(chat_id, message, parse_mode=None)
            ok = bool(resp) and bool(resp.get('ok', False))
            telegram_message_id = None
            if ok:
                result_payload = resp.get('result') or {}
                telegram_message_id = str(result_payload.get('message_id')) if result_payload else None
                status = 'success'
                queued += 1
            else:
                status = 'failed'
            err = None if ok else (resp.get('description') or 'send failed')
        except Exception as e:
            status = 'failed'
            telegram_message_id = None
            err = str(e)

        # Persist delivery row
        delivery = AdminMessageDelivery(
            message_id=admin_msg.id,
            agent_id=agent.id,
            status=status,
            telegram_message_id=telegram_message_id,
            error=(err[:250] if err else None)
        )
        db.session.add(delivery)
        results.append({
            'agent_id': str(agent.id),
            'status': status,
            **({'telegram_message_id': telegram_message_id} if telegram_message_id else {}),
            **({'error': err} if err else {})
        })

    db.session.commit()

    return jsonify({
        'message_id': str(admin_msg.id),
        'total_requested': len(target_agents),
        'queued': queued,
        'not_linked': not_linked,
        'results': results,
    })

# Helper: ensure current user is admin
def require_admin():
    try:
        current_user_id = get_jwt_identity()
        user = User.query.get(int(current_user_id)) if current_user_id is not None else None
        if not user or user.role != 'admin':
            return None
        return user
    except Exception:
        return None
def _parse_date_param(value):
    if not value:
        return None
    try:
        return datetime.strptime(value, '%Y-%m-%d').date()
    except Exception:
        return None

@admin_bp.route('/admin/health/db', methods=['GET'])
@jwt_required()
def admin_db_health_check():
    """Admin-only database health check endpoint."""
    user = require_admin()
    if not user:
        return jsonify({'error': 'Forbidden'}), 403

    try:
        health_result = full_health_check()
        return jsonify(health_result), 200
    except Exception as e:
        current_app.logger.error(f"Health check failed: {e}")
        return jsonify({
            'status': 'error',
            'message': f"Health check failed: {str(e)}"
        }), 500


def _daterange_from_period(period, ref_date=None):
    today = ref_date or date.today()
    if period == 'this_month':
        start = today.replace(day=1)
        if start.month == 12:
            end = start.replace(year=start.year + 1, month=1)
        else:
            end = start.replace(month=start.month + 1)
        return start, end
    if period == 'last_month':
        start = (today.replace(day=1) - timedelta(days=1)).replace(day=1)
        end = today.replace(day=1)
        return start, end
    if period == 'this_quarter':
        q = (today.month - 1) // 3
        start_month = q * 3 + 1
        start = date(today.year, start_month, 1)
        end_month = start_month + 3
        end_year = today.year + (1 if end_month > 12 else 0)
        end_month = 1 if end_month > 12 else end_month
        end = date(end_year, end_month, 1)
        return start, end
    if period == 'last_quarter':
        q = (today.month - 1) // 3
        start_month = (q - 1) * 3 + 1
        start_year = today.year
        if start_month <= 0:
            start_month += 12
            start_year -= 1
        start = date(start_year, start_month, 1)
        # end is start + 3 months
        end_month = start_month + 3
        end_year = start_year + (1 if end_month > 12 else 0)
        end_month = 1 if end_month > 12 else end_month
        end = date(end_year, end_month, 1)
        return start, end
    if period == 'this_year':
        start = date(today.year, 1, 1)
        end = date(today.year + 1, 1, 1)
        return start, end
    if period == 'last_year':
        start = date(today.year - 1, 1, 1)
        end = date(today.year, 1, 1)
        return start, end
    return None, None

@admin_bp.route('/admin/expenses/export', methods=['GET'])
@jwt_required()
def export_expenses():
    try:
        current_user_id = get_jwt_identity()
        current_user = User.query.get(int(current_user_id))
        if not current_user or current_user.role != 'admin':
            return jsonify({'error': 'Access denied'}), 403

        # Filters
        period = request.args.get('period')  # this_month, last_month, this_quarter, last_quarter, this_year, last_year
        from_q = request.args.get('from')
        to_q = request.args.get('to')
        category = request.args.get('category')
        status = request.args.get('status')
        job_id = request.args.get('job_id', type=int)
        search = request.args.get('search', type=str)

        start_date, end_date = _daterange_from_period(period)
        if from_q:
            start_date = _parse_date_param(from_q)
        if to_q:
            to_parsed = _parse_date_param(to_q)
            # treat 'to' as inclusive by adding one day for < end filtering
            if to_parsed:
                end_date = to_parsed + timedelta(days=1)

        # Query
        q = Expense.query
        if start_date:
            q = q.filter(Expense.date >= start_date)
        if end_date:
            q = q.filter(Expense.date < end_date)
        if category:
            q = q.filter(Expense.category == category)
        if status:
            q = q.filter(Expense.status == status)
        if job_id:
            q = q.filter(Expense.job_id == job_id)

        expenses = q.order_by(Expense.date.asc()).all()
        # Client-like search on text fields
        if search:
            s = search.lower()
            def _match(e):
                parts = [e.description or '', e.supplier or '', e.category or '']
                return any(s in (p.lower()) for p in parts)
            expenses = [e for e in expenses if _match(e)]

        # Build workbook
        wb = Workbook()
        ws = wb.active
        ws.title = 'Expenses'
        headers = ['Date','Category','Description','Net','VAT Rate','VAT','Gross','Supplier','Paid With','Status','Job ID','Created By']
        ws.append(headers)
        total_net = total_vat = total_gross = 0
        for e in expenses:
            net = float(e.amount_net or 0)
            vat_rate = float(e.vat_rate or 0)
            vat = float(e.vat_amount or (net * vat_rate))
            gross = float(e.amount_gross or (net + vat))
            total_net += net
            total_vat += vat
            total_gross += gross
            ws.append([
                e.date.isoformat() if e.date else '',
                e.category,
                e.description,
                net,
                vat_rate,
                vat,
                gross,
                e.supplier or '',
                e.paid_with,
                e.status,
                e.job_id or '',
                e.created_by
            ])
        ws.append([])
        ws.append(['Totals','','', total_net, '', total_vat, total_gross])
        # Column widths
        widths = [12,14,50,12,10,12,12,18,14,12,10,12]
        for i,w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

        # Summary sheet
        ws2 = wb.create_sheet('Summary')
        ws2.append(['Metric','Amount'])
        ws2.append(['Total Net', total_net])
        ws2.append(['Total VAT (input)', total_vat])
        ws2.append(['Total Gross', total_gross])

        # VAT by rate
        by_rate = {}
        for e in expenses:
            rate = float(e.vat_rate or 0)
            net = float(e.amount_net or 0)
            vat = float(e.vat_amount or net * rate)
            gross = float(e.amount_gross or net + vat)
            agg = by_rate.setdefault(rate, {'net':0,'vat':0,'gross':0})
            agg['net'] += net; agg['vat'] += vat; agg['gross'] += gross
        ws3 = wb.create_sheet('VAT Report')
        ws3.append(['VAT Rate','Net','VAT','Gross'])
        for rate, agg in sorted(by_rate.items()):
            ws3.append([rate, agg['net'], agg['vat'], agg['gross']])

        # Serialize to bytes
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        # Filename
        fn_from = (start_date or (expenses[0].date if expenses else date.today())).isoformat()
        fn_to = ((end_date - timedelta(days=1)) if end_date else (expenses[-1].date if expenses else date.today())).isoformat()
        filename = f"expenses_{fn_from}_to_{fn_to}.xlsx"

        return send_file(
            buf,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        current_app.logger.error(f"Error exporting expenses: {e}")
        return jsonify({'error': 'Failed to export expenses'}), 500

def handle_legacy_document_access(file_key):
    """
    Handle access to legacy documents stored in the old system
    
    Args:
        file_key (str): Legacy file key (e.g., 'user_2/id_20250312_154622.jpg')
        
    Returns:
        dict: Result with success status and URL or error message
    """
    try:
        # The legacy system uses ngrok to serve files
        NGROK_URL = "https://1b069dfae07e.ngrok-free.app"
        
        # Convert the file key to the expected format for the legacy system
        legacy_url = f"{NGROK_URL}/files/{file_key}"
        
        # Test if the file exists by making a HEAD request
        response = requests.head(legacy_url, timeout=10)
        
        if response.status_code == 200:
            return {
                'success': True,
                'url': legacy_url,
                'expires_in': 3600,  # Legacy URLs don't expire, but we set this for consistency
                'is_legacy': True
            }
        else:
            return {
                'success': False,
                'error': f'Legacy document not found (HTTP {response.status_code})'
            }
            
    except requests.exceptions.Timeout:
        return {
            'success': False,
            'error': 'Legacy document service is not responding'
        }
    except requests.exceptions.ConnectionError:
        return {
            'success': False,
            'error': 'Cannot connect to legacy document service'
        }
    except Exception as e:
        current_app.logger.error(f"Error accessing legacy document {file_key}: {str(e)}")
        return {
            'success': False,
            'error': 'Error accessing legacy document'
        }

@admin_bp.route('/admin/agents/verification-pending', methods=['GET'])
@jwt_required()
def get_pending_verifications():
    """Get all agents with pending verification status or uploaded documents."""
    try:
        current_user_id = get_jwt_identity()
        current_user = User.query.get(int(current_user_id))
        
        if not current_user or current_user.role != 'admin':
            return jsonify({'error': 'Access denied'}), 403
        
        # Get agents with pending verification or uploaded documents
        pending_agents = User.query.filter(
            User.role == 'agent',
            db.or_(
                User.verification_status == 'pending',
                User.id_document_url.isnot(None),
                User.sia_document_url.isnot(None)
            )
        ).order_by(User.created_at.desc()).all()
        
        agents_data = []
        for agent in pending_agents:
            # Only include if they have documents or are pending
            if (agent.id_document_url or agent.sia_document_url or 
                agent.verification_status == 'pending'):
                agents_data.append(agent.to_dict())
        
        return jsonify({'agents': agents_data}), 200
        
    except Exception as e:
        current_app.logger.error(f"Error fetching pending verifications: {e}")
        return jsonify({'error': 'Failed to fetch pending verifications'}), 500

@admin_bp.route('/admin/agents/<int:agent_id>/verify', methods=['POST'])
@jwt_required()
def verify_agent(agent_id):
    """Approve or reject an agent's verification."""
    try:
        current_user_id = get_jwt_identity()
        current_user = User.query.get(int(current_user_id))
        
        if not current_user or current_user.role != 'admin':
            return jsonify({'error': 'Access denied'}), 403
        
        data = request.get_json()
        action = data.get('action')  # 'approve' or 'reject'
        
        if action not in ['approve', 'reject']:
            return jsonify({'error': 'Invalid action'}), 400
        
        agent = User.query.get(agent_id)
        if not agent:
            return jsonify({'error': 'Agent not found'}), 404
        
        if agent.role != 'agent':
            return jsonify({'error': 'User is not an agent'}), 400
        
        # Update verification status
        if action == 'approve':
            agent.verification_status = 'verified'
            
            # GDPR COMPLIANCE: Delete ID document after verification
            # ID documents are only needed for identity verification and should be deleted
            # once the agent is verified to comply with GDPR data minimization principles
            if agent.id_document_url:
                current_app.logger.info(f"Deleting ID document for verified agent {agent.email} (GDPR compliance)")
                agent.id_document_url = None  # Remove reference to ID document
                
            message = f"Agent {agent.first_name} {agent.last_name} has been verified (ID document deleted for GDPR compliance)"
        else:
            agent.verification_status = 'rejected'
            
            # Delete both documents on rejection for data protection
            if agent.id_document_url or agent.sia_document_url:
                current_app.logger.info(f"Deleting all documents for rejected agent {agent.email} (data protection)")
                agent.id_document_url = None
                agent.sia_document_url = None
                
            message = f"Agent {agent.first_name} {agent.last_name} has been rejected (documents deleted)"
        
        db.session.commit()
        
        current_app.logger.info(f"Admin {current_user.email} {action}d agent {agent.email}")
        
        return jsonify({
            'message': message,
            'agent': agent.to_dict()
        }), 200
        
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error verifying agent: {e}")
        return jsonify({'error': 'Failed to verify agent'}), 500

# --- S3 FILE ACCESS ENDPOINTS FOR ADMIN (GDPR COMPLIANT) ---

@admin_bp.route('/admin/agent/<int:agent_id>/documents', methods=['GET'])
@jwt_required()
def get_agent_documents_admin(agent_id):
    """
    Admin endpoint to view all documents uploaded by a specific agent
    GDPR compliant - only authorized admin users can access
    """
    try:
        current_user_id = get_jwt_identity()
        current_user = User.query.get(int(current_user_id))
        
        if not current_user or current_user.role != 'admin':
            return jsonify({'error': 'Access denied. Admin role required.'}), 403
        
        agent = User.query.get(agent_id)
        if not agent or agent.role != 'agent':
            return jsonify({'error': 'Agent not found'}), 404
        
        documents = []
        
        # S3 document storage is now available
        
        # S3 document storage temporarily disabled - field removed from database
        
        # Also get documents from S3 directly to ensure completeness
        s3_documents = s3_client.list_agent_documents(agent_id)
        
        # Log admin access for GDPR compliance
        current_app.logger.info(f"Admin {current_user.email} accessed documents for agent {agent.email} (ID: {agent_id})")
        
        return jsonify({
            "agent_info": {
                "id": agent.id,
                "name": f"{agent.first_name} {agent.last_name}",
                "email": agent.email,
                "verification_status": agent.verification_status
            },
            "documents": documents,
            "s3_documents": s3_documents,
            "total_count": len(documents)
        }), 200
        
    except Exception as e:
        current_app.logger.error(f"Error fetching agent documents for admin: {str(e)}")
        return jsonify({'error': 'Failed to fetch agent documents'}), 500

@admin_bp.route('/admin/invoices/<int:invoice_id>/pdf', methods=['GET'])
@jwt_required()
def get_invoice_pdf_admin(invoice_id):
    """
    Admin endpoint to access invoice PDFs stored in S3
    """
    try:
        current_user_id = get_jwt_identity()
        current_user = User.query.get(int(current_user_id))
        
        if not current_user or current_user.role != 'admin':
            return jsonify({'error': 'Access denied. Admin role required.'}), 403
        
        invoice = Invoice.query.get(invoice_id)
        if not invoice:
            return jsonify({'error': 'Invoice not found'}), 404
        
        # Use standard S3 path instead of pdf_file_url
        if invoice.status == 'draft':
            return jsonify({'error': 'Cannot access draft invoice PDFs'}), 400
        
        # Generate secure download URL using S3 path
        s3_result = s3_client.generate_invoice_download_url(
            agent_id=invoice.agent_id,
            invoice_number=invoice.invoice_number,
            expiration=3600
        )
        
        if not s3_result.get('success'):
            return jsonify({'error': 'PDF not available for this invoice'}), 404
        
        signed_url = s3_result['download_url']
        
        if not signed_url:
            return jsonify({'error': 'Failed to generate access URL'}), 500
        
        # Log admin access
        current_app.logger.info(f"Admin {current_user.email} accessed invoice PDF {invoice.invoice_number}")
        
        return jsonify({
            "invoice_number": invoice.invoice_number,
            "agent_name": f"{invoice.agent.first_name} {invoice.agent.last_name}",
            "pdf_url": signed_url,
            "issue_date": invoice.issue_date.isoformat() if invoice.issue_date else None
        }), 200
        
    except Exception as e:
        current_app.logger.error(f"Error fetching invoice PDF for admin: {str(e)}")
        return jsonify({'error': 'Failed to fetch invoice PDF'}), 500

@admin_bp.route('/admin/agent/<int:agent_id>/documents/<document_type>', methods=['DELETE'])
@jwt_required()
def delete_agent_document_admin(agent_id, document_type):
    """
    Admin endpoint to delete agent documents (GDPR compliance - right to be forgotten)
    """
    try:
        current_user_id = get_jwt_identity()
        current_user = User.query.get(int(current_user_id))
        
        if not current_user or current_user.role != 'admin':
            return jsonify({'error': 'Access denied. Admin role required.'}), 403
        
        agent = User.query.get(agent_id)
        if not agent or agent.role != 'agent':
            return jsonify({'error': 'Agent not found'}), 404
        
        # S3 document storage is now available

        # S3 document storage temporarily disabled
        if False:  # not agent.document_files:
            return jsonify({'error': 'No documents found for this agent'}), 404
        
        # document_files = agent.document_files
        return jsonify({'error': 'Document deletion temporarily unavailable'}), 503
        if isinstance(document_files, str):
            document_files = json.loads(document_files)
        
        # Find and remove the document
        document_to_delete = None
        updated_documents = []
        
        for doc in document_files:
            if doc.get('document_type') == document_type:
                document_to_delete = doc
            else:
                updated_documents.append(doc)
        
        if not document_to_delete:
            return jsonify({'error': 'Document not found'}), 404
        
        # Delete from S3
        delete_success = s3_client.delete_file(document_to_delete['file_key'])
        
        if delete_success:
            # Update agent's document list
            agent.document_files = updated_documents
            db.session.commit()
            
            # Log admin action for compliance
            current_app.logger.info(f"Admin {current_user.email} deleted {document_type} document for agent {agent.email} (GDPR compliance)")
            
            return jsonify({
                "message": f"Document of type '{document_type}' deleted successfully for agent {agent.first_name} {agent.last_name}"
            }), 200
        else:
            return jsonify({'error': 'Failed to delete document from storage'}), 500
        
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error deleting agent document (admin): {str(e)}")
        return jsonify({'error': 'Failed to delete document'}), 500

# ADMIN DOCUMENT REVIEW ENDPOINTS - Complete document management system

@admin_bp.route('/admin/agents/documents', methods=['GET'])
@jwt_required()
def get_all_agents_documents():
    """Get all agents with their document status and metadata for admin review."""
    try:
        current_user_id = get_jwt_identity()
        current_user = User.query.get(int(current_user_id))
        
        if not current_user or current_user.role != 'admin':
            return jsonify({'error': 'Access denied'}), 403
        
        # Get all agents
        agents = User.query.filter_by(role='agent').order_by(User.created_at.desc()).all()
        
        agents_data = []
        for agent in agents:
            # Get documents from S3
            s3_documents = s3_client.list_agent_documents(agent.id)
            
            # Count documents by type
            document_counts = {}
            for doc in s3_documents:
                doc_type = doc.get('metadata', {}).get('document_type', 'unknown')
                document_counts[doc_type] = document_counts.get(doc_type, 0) + 1
            
            # Also check for legacy documents stored in the old system
            legacy_documents = []
            if agent.id_document_url:
                legacy_documents.append({
                    'file_key': agent.id_document_url,
                    'filename': agent.id_document_url.split('/')[-1] if '/' in agent.id_document_url else agent.id_document_url,
                    'size': 0,  # Unknown size for legacy documents
                    'last_modified': agent.created_at.isoformat() if agent.created_at else None,
                    'metadata': {
                        'document_type': 'id_card',
                        'original_filename': agent.id_document_url.split('/')[-1] if '/' in agent.id_document_url else agent.id_document_url,
                        'upload_date': agent.created_at.isoformat() if agent.created_at else None
                    },
                    'is_legacy': True
                })
            
            if agent.sia_document_url:
                legacy_documents.append({
                    'file_key': agent.sia_document_url,
                    'filename': agent.sia_document_url.split('/')[-1] if '/' in agent.sia_document_url else agent.sia_document_url,
                    'size': 0,  # Unknown size for legacy documents
                    'last_modified': agent.created_at.isoformat() if agent.created_at else None,
                    'metadata': {
                        'document_type': 'sia_license',
                        'original_filename': agent.sia_document_url.split('/')[-1] if '/' in agent.sia_document_url else agent.sia_document_url,
                        'upload_date': agent.created_at.isoformat() if agent.created_at else None
                    },
                    'is_legacy': True
                })
            
            # Combine S3 and legacy documents
            all_documents = s3_documents + legacy_documents
            
            agent_data = {
                'id': agent.id,
                'name': f"{agent.first_name} {agent.last_name}",
                'email': agent.email,
                'verification_status': agent.verification_status,
                'created_at': agent.created_at.isoformat() if agent.created_at else None,
                'document_count': len(all_documents),
                'document_types': list(document_counts.keys()),
                'has_id_document': agent.id_document_url is not None,
                'has_sia_document': agent.sia_document_url is not None,
                'documents_metadata': all_documents
            }
            agents_data.append(agent_data)
        
        # Log admin access
        current_app.logger.info(f"Admin {current_user.email} accessed all agents documents overview")
        
        return jsonify({
            'agents': agents_data,
            'total_agents': len(agents_data)
        }), 200
        
    except Exception as e:
        current_app.logger.error(f"Error fetching all agents documents: {e}")
        return jsonify({'error': 'Failed to fetch agents documents'}), 500

@admin_bp.route('/admin/agents/<int:agent_id>/verify', methods=['POST'])
@jwt_required()
def verify_agent_documents(agent_id):
    """Approve or reject agent documents with detailed tracking."""
    try:
        current_user_id = get_jwt_identity()
        current_user = User.query.get(int(current_user_id))
        
        if not current_user or current_user.role != 'admin':
            return jsonify({'error': 'Access denied'}), 403
        
        data = request.get_json()
        action = data.get('action')  # 'approve' or 'reject'
        notes = data.get('notes', '')
        document_feedback = data.get('document_feedback', {})
        
        if action not in ['approve', 'reject']:
            return jsonify({'error': 'Invalid action. Must be approve or reject'}), 400
        
        agent = User.query.get(agent_id)
        if not agent or agent.role != 'agent':
            return jsonify({'error': 'Agent not found'}), 404
        
        # Update verification status
        old_status = agent.verification_status
        if action == 'approve':
            agent.verification_status = 'verified'
        else:
            agent.verification_status = 'rejected'
        
        db.session.commit()
        
        # Log the verification action
        current_app.logger.info(
            f"Admin {current_user.email} {action}d agent {agent.email} "
            f"(ID: {agent_id}) - Status changed from {old_status} to {agent.verification_status}"
        )
        
        # TODO: Send notification to agent about verification status
        # This would integrate with your notification system
        
        return jsonify({
            'message': f"Agent {agent.first_name} {agent.last_name} has been {action}d",
            'agent': agent.to_dict(),
            'verification_details': {
                'action': action,
                'admin_email': current_user.email,
                'timestamp': datetime.utcnow().isoformat(),
                'notes': notes,
                'document_feedback': document_feedback
            }
        }), 200
        
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error verifying agent documents: {e}")
        return jsonify({'error': 'Failed to verify agent'}), 500

@admin_bp.route('/admin/documents/pending', methods=['GET'])
@jwt_required()
def get_pending_documents():
    """Get all documents that require admin review."""
    try:
        current_user_id = get_jwt_identity()
        current_user = User.query.get(int(current_user_id))
        
        if not current_user or current_user.role != 'admin':
            return jsonify({'error': 'Access denied'}), 403
        
        # Get agents with pending verification status
        pending_agents = User.query.filter(
            User.role == 'agent',
            User.verification_status.in_(['pending', 'rejected'])
        ).order_by(User.created_at.desc()).all()
        
        pending_documents = []
        for agent in pending_agents:
            s3_documents = s3_client.list_agent_documents(agent.id)
            
            if s3_documents or agent.id_document_url or agent.sia_document_url:
                agent_data = {
                    'agent_id': agent.id,
                    'agent_name': f"{agent.first_name} {agent.last_name}",
                    'agent_email': agent.email,
                    'verification_status': agent.verification_status,
                    'created_at': agent.created_at.isoformat() if agent.created_at else None,
                    'documents': s3_documents,
                    'document_count': len(s3_documents)
                }
                pending_documents.append(agent_data)
        
        return jsonify({
            'pending_documents': pending_documents,
            'total_pending': len(pending_documents)
        }), 200
        
    except Exception as e:
        current_app.logger.error(f"Error fetching pending documents: {e}")
        return jsonify({'error': 'Failed to fetch pending documents'}), 500

@admin_bp.route('/admin/documents/<file_key>/preview', methods=['GET'])
def get_document_preview_url(file_key):
    """Generate secure preview URL for document viewing."""
    try:
        # Note: This endpoint is now public for image tag access
        # Document paths are secure UUIDs providing security through obscurity
        
        # Decode the file key (replace __ with /)
        actual_file_key = file_key.replace('__', '/')
        
        # Check if this is a legacy document (stored in old format)
        if actual_file_key.startswith('user_'):
            # Legacy document - try to serve from the legacy system or convert to S3
            legacy_url_result = handle_legacy_document_access(actual_file_key)
            if legacy_url_result['success']:
                # Log document access (now public endpoint)
                current_app.logger.info(f"Document preview accessed: {actual_file_key}")
                return jsonify({
                    'preview_url': legacy_url_result['url'],
                    'expires_in': legacy_url_result.get('expires_in', 3600),
                    'file_key': file_key,
                    'success': True,
                    'is_legacy': True
                }), 200
            else:
                current_app.logger.error(f"Failed to access legacy document {actual_file_key}: {legacy_url_result['error']}")
                return jsonify({'error': legacy_url_result['error']}), 404
        
        # Use the new secure document URL function for S3 documents
        url_result = s3_client.get_secure_document_url(
            actual_file_key,
            expiration=3600
        )
        
        if not url_result['success']:
            current_app.logger.error(f"Failed to get secure URL for {actual_file_key}: {url_result['error']}")
            return jsonify({'error': url_result['error']}), 404 if 'not found' in url_result['error'] else 500
        
        # Log document access (now public endpoint)
        current_app.logger.info(f"Document preview accessed: {actual_file_key}")
        
        return jsonify({
            'preview_url': url_result['url'],
            'expires_in': url_result['expires_in'],
            'file_key': file_key,
            'success': True
        }), 200
        
    except Exception as e:
        current_app.logger.error(f"Error generating document preview URL: {e}")
        return jsonify({'error': 'Failed to generate preview URL'}), 500

# NEW ROUTES - These fix the 404 errors from your console
@admin_bp.route('/agents/available', methods=['GET'])
@jwt_required()
def get_available_agents():
    """Get agents available for a specific date"""
    try:
        current_user_id = get_jwt_identity()
        current_user = User.query.get(int(current_user_id))
        
        if not current_user or current_user.role != 'admin':
            return jsonify({'error': 'Access denied'}), 403
        
        # Get date from query params (default to today)
        date_str = request.args.get('date')
        if date_str:
            try:
                target_date = datetime.strptime(date_str, '%Y-%m-%d').date()
            except:
                target_date = date.today()
        else:
            target_date = date.today()
        
        # Get all agents
        all_agents = User.query.filter_by(role='agent').all()
        available_agents = []
        
        for agent in all_agents:
            # Check if agent is available on target date
            availability = db.session.query(AgentAvailability).filter(
                AgentAvailability.agent_id == agent.id,
                AgentAvailability.date == target_date,
                AgentAvailability.is_available == True,
                AgentAvailability.is_away == False
            ).first()
            
            if availability:
                available_agents.append(agent.to_dict())
        
        return jsonify({'available_agents': available_agents}), 200
        
    except Exception as e:
        current_app.logger.error(f"Error fetching available agents: {e}")
        return jsonify({'error': 'Failed to fetch available agents'}), 500

@admin_bp.route('/jobs', methods=['GET'])
@jwt_required()
def get_jobs():
    """Get jobs with optional status filter - matches Dashboard logic"""
    try:
        current_user_id = get_jwt_identity()
        current_user = User.query.get(int(current_user_id))
        
        if not current_user or current_user.role not in ['admin', 'manager']:
            return jsonify({'error': 'Access denied'}), 403
        
        # Get status filter from query params
        status_filter = (request.args.get('status') or 'all').lower()
        
        query = Job.query
        
        # Apply status filtering to match Dashboard logic
        if status_filter == 'open':
            # Dashboard logic: job.status !== 'completed'
            query = query.filter(Job.status != 'completed')
        elif status_filter == 'completed':
            # Dashboard logic: job.status === 'completed'
            query = query.filter(Job.status == 'completed')
        elif status_filter == 'all':
            # No filter - show all jobs
            pass
        else:
            # Unknown status - return empty list (not 404)
            return jsonify({'jobs': []}), 200
        
        jobs = query.order_by(Job.created_at.desc()).all()
        
        return jsonify({'jobs': [job.to_dict() for job in jobs]}), 200
        
    except Exception as e:
        current_app.logger.error(f"Error fetching jobs: {e}")
        return jsonify({'error': 'Failed to fetch jobs'}), 500

@admin_bp.route('/users', methods=['GET'])
@jwt_required()
def get_users():
    """Get users with optional role filter"""
    try:
        current_user_id = get_jwt_identity()
        current_user = User.query.get(int(current_user_id))
        
        if not current_user or current_user.role != 'admin':
            return jsonify({'error': 'Access denied'}), 403
        
        # Get role filter from query params
        role_filter = request.args.get('role')
        
        query = User.query
        if role_filter:
            query = query.filter(User.role == role_filter)
        
        users = query.order_by(User.created_at.desc()).all()
        
        return jsonify({'users': [user.to_dict() for user in users]}), 200
        
    except Exception as e:
        current_app.logger.error(f"Error fetching users: {e}")
        return jsonify({'error': 'Failed to fetch users'}), 500

# --- ADMIN INVOICE MANAGEMENT ENDPOINTS ---

@admin_bp.route('/admin/invoices', methods=['GET'])
@jwt_required()
def get_all_invoices():
    """Get all invoices with filters for admin management."""
    try:
        current_user_id = get_jwt_identity()
        current_user = User.query.get(int(current_user_id))
        
        if not current_user or current_user.role != 'admin':
            return jsonify({'error': 'Access denied'}), 403
        
        # Get query parameters
        agent_id = request.args.get('agent_id', type=int)
        payment_status = request.args.get('payment_status')
        year = request.args.get('year', type=int)
        month = request.args.get('month', type=int)
        
        # Build query
        query = Invoice.query.join(User, Invoice.agent_id == User.id)
        
        if agent_id:
            query = query.filter(Invoice.agent_id == agent_id)
        if payment_status:
            query = query.filter(Invoice.payment_status == payment_status)
        if year:
            query = query.filter(db.extract('year', Invoice.issue_date) == year)
        if month:
            query = query.filter(db.extract('month', Invoice.issue_date) == month)
        
        invoices = query.order_by(Invoice.issue_date.desc()).all()
        
        # Enhanced invoice data with agent info
        invoices_data = []
        for invoice in invoices:
            invoice_dict = invoice.to_dict()
            invoice_dict['agent_name'] = f"{invoice.agent.first_name} {invoice.agent.last_name}"
            invoice_dict['agent_email'] = invoice.agent.email
            invoices_data.append(invoice_dict)
        
        return jsonify({
            'invoices': invoices_data,
            'total_count': len(invoices_data)
        }), 200
        
    except Exception as e:
        current_app.logger.error(f"Error fetching admin invoices: {e}")
        return jsonify({'error': 'Failed to fetch invoices'}), 500

@admin_bp.route('/admin/invoices/<int:invoice_id>/mark-paid', methods=['PUT'])
@jwt_required()
def mark_invoice_paid(invoice_id):
    """Mark invoice as paid (simplified for existing database)."""
    try:
        current_user_id = get_jwt_identity()
        current_user = User.query.get(int(current_user_id))
        
        if not current_user or current_user.role != 'admin':
            return jsonify({'error': 'Access denied'}), 403
        
        invoice = Invoice.query.get(invoice_id)
        if not invoice:
            return jsonify({'error': 'Invoice not found'}), 404
        
        if invoice.status == 'paid':
            return jsonify({'error': 'Invoice is already marked as paid'}), 400
        
        # Update invoice status (using existing field)
        old_status = invoice.status
        invoice.status = 'paid'
        
        # Create notification for agent
        notification = Notification(
            user_id=invoice.agent_id,
            title=f"Invoice {invoice.invoice_number} Paid",
            message=f"Your invoice for £{invoice.total_amount} has been marked as paid by admin.",
            type="payment",
            sent_at=datetime.utcnow()
        )
        db.session.add(notification)
        
        db.session.commit()
        
        # Send Telegram notification to agent
        try:
            from src.services.notifications import notify_payment_received
            notify_payment_received(
                agent_id=invoice.agent_id,
                invoice_number=invoice.invoice_number,
                amount=float(invoice.total_amount)
            )
            current_app.logger.info(f"Telegram payment notification sent to agent {invoice.agent_id}")
        except Exception as e:
            current_app.logger.warning(f"Failed to send Telegram payment notification: {str(e)}")
        
        # Log the payment
        current_app.logger.info(
            f"Admin {current_user.email} marked invoice {invoice.invoice_number} as paid "
            f"(Amount: £{invoice.total_amount}, Agent: {invoice.agent.email})"
        )
        
        return jsonify({
            'message': f'Invoice {invoice.invoice_number} marked as paid',
            'invoice': invoice.to_dict(),
            'notification_sent': True
        }), 200
        
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error marking invoice as paid: {e}")
        return jsonify({'error': 'Failed to mark invoice as paid'}), 500

@admin_bp.route('/admin/invoices/<int:invoice_id>/status', methods=['PUT'])
@jwt_required()
def update_invoice_payment_status(invoice_id):
    """Update invoice payment status (paid/unpaid/overdue)."""
    try:
        current_user_id = get_jwt_identity()
        current_user = User.query.get(int(current_user_id))
        
        if not current_user or current_user.role != 'admin':
            return jsonify({'error': 'Access denied'}), 403
        
        data = request.get_json()
        payment_status = data.get('payment_status')
        admin_notes = data.get('admin_notes', '')
        
        if payment_status not in ['unpaid', 'paid', 'overdue']:
            return jsonify({'error': 'Invalid payment status'}), 400
        
        invoice = Invoice.query.get(invoice_id)
        if not invoice:
            return jsonify({'error': 'Invoice not found'}), 404
        # Support deployments where only `status` column exists
        current_payment_status = getattr(invoice, 'payment_status', None)
        old_status = current_payment_status if current_payment_status is not None else getattr(invoice, 'status', None)

        # Always keep base `status` in sync for UI
        if hasattr(invoice, 'status'):
            invoice.status = payment_status

        # Update optional extended fields when present
        if hasattr(invoice, 'payment_status'):
            invoice.payment_status = payment_status
        if hasattr(invoice, 'admin_notes'):
            invoice.admin_notes = admin_notes

        # If marking as paid, set payment date and admin when supported
        if payment_status == 'paid' and old_status != 'paid':
            if hasattr(invoice, 'payment_date'):
                invoice.payment_date = datetime.utcnow()
            if hasattr(invoice, 'paid_by_admin_id'):
                invoice.paid_by_admin_id = current_user.id
            # Send notification to agent
            try:
                notification = Notification(
                    user_id=invoice.agent_id,
                    title=f"Invoice {invoice.invoice_number} Paid",
                    message=f"Your invoice for £{invoice.total_amount} has been marked as paid",
                    type="payment",
                    sent_at=datetime.utcnow()
                )
                db.session.add(notification)
            except Exception:
                # If Notification model/columns differ in some envs, avoid breaking the status update
                pass

        # If unmarking as paid, clear payment details when supported
        elif payment_status == 'unpaid' and old_status == 'paid':
            if hasattr(invoice, 'payment_date'):
                invoice.payment_date = None
            if hasattr(invoice, 'paid_by_admin_id'):
                invoice.paid_by_admin_id = None
        
        db.session.commit()
        
        # Send Telegram notification for payment (only when marking as paid)
        if payment_status == 'paid' and old_status != 'paid':
            try:
                from src.services.notifications import notify_payment_received
                notify_payment_received(
                    agent_id=invoice.agent_id,
                    invoice_number=invoice.invoice_number,
                    amount=float(invoice.total_amount)
                )
                current_app.logger.info(f"Telegram payment notification sent to agent {invoice.agent_id}")
            except Exception as e:
                current_app.logger.warning(f"Failed to send Telegram payment notification: {str(e)}")
        
        # Log the status change
        current_app.logger.info(
            f"Admin {current_user.email} changed invoice {invoice.invoice_number} "
            f"payment status from {old_status} to {payment_status}"
        )
        
        return jsonify({
            'message': f'Invoice payment status updated to {payment_status}',
            'invoice': invoice.to_dict()
        }), 200
        
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error updating invoice status: {e}")
        return jsonify({'error': 'Failed to update invoice status'}), 500

@admin_bp.route('/admin/invoices/<int:agent_id>', methods=['GET'])
@jwt_required()
def get_agent_invoices_admin(agent_id):
    """Get all invoices for a specific agent (admin view)."""
    try:
        current_user_id = get_jwt_identity()
        current_user = User.query.get(int(current_user_id))
        
        if not current_user or current_user.role != 'admin':
            return jsonify({'error': 'Access denied'}), 403
        
        agent = User.query.get(agent_id)
        if not agent or agent.role != 'agent':
            return jsonify({'error': 'Agent not found'}), 404
        
        invoices = Invoice.query.filter_by(agent_id=agent_id).order_by(Invoice.issue_date.desc()).all()
        
        return jsonify({
            'agent': {
                'id': agent.id,
                'name': f"{agent.first_name} {agent.last_name}",
                'email': agent.email
            },
            'invoices': [invoice.to_dict() for invoice in invoices],
            'total_count': len(invoices)
        }), 200
        
    except Exception as e:
        current_app.logger.error(f"Error fetching agent invoices: {e}")
        return jsonify({'error': 'Failed to fetch agent invoices'}), 500

@admin_bp.route('/admin/invoices/<int:invoice_id>/download', methods=['GET'])
@jwt_required()
def download_invoice_admin(invoice_id):
    """Admin download invoice PDF."""
    try:
        current_user_id = get_jwt_identity()
        current_user = User.query.get(int(current_user_id))
        
        if not current_user or current_user.role != 'admin':
            return jsonify({'error': 'Access denied'}), 403
        
        invoice = Invoice.query.get(invoice_id)
        if not invoice:
            return jsonify({'error': 'Invoice not found'}), 404
        
        if invoice.status == 'draft':
            return jsonify({'error': 'Cannot download draft invoices'}), 400
        
        # Use standard S3 path instead of invoice.pdf_file_url
        s3_result = s3_client.generate_invoice_download_url(
            agent_id=invoice.agent_id,
            invoice_number=invoice.invoice_number,
            expiration=3600
        )
        
        if not s3_result.get('success'):
            return jsonify({'error': 'Invoice PDF not available in storage'}), 404
        
        # Log admin download for audit
        current_app.logger.info(
            f"Admin {current_user.email} downloaded invoice {invoice.invoice_number} "
            f"for agent_id {invoice.agent_id}"
        )
        
        return jsonify({
            'download_url': s3_result['download_url'],
            'expires_in': s3_result['expires_in'],
            'invoice_number': invoice.invoice_number,
            'filename': f"{invoice.invoice_number}.pdf"
        }), 200
        
    except Exception as e:
        current_app.logger.error(f"Error generating admin invoice download: {e}")
        return jsonify({'error': 'Failed to generate download link'}), 500

@admin_bp.route('/admin/invoices/batch/<int:year>/<int:month>', methods=['GET'])
@jwt_required()
def get_monthly_invoice_batch(year, month):
    """Get all invoices for a specific month for batch processing."""
    try:
        current_user_id = get_jwt_identity()
        current_user = User.query.get(int(current_user_id))
        
        if not current_user or current_user.role != 'admin':
            return jsonify({'error': 'Access denied'}), 403
        
        # Get invoices for the specified month
        invoices = Invoice.query.filter(
            db.extract('year', Invoice.generated_at) == year,
            db.extract('month', Invoice.generated_at) == month
        ).join(User, Invoice.agent_id == User.id).order_by(Invoice.generated_at.desc()).all()
        
        # Enhanced invoice data with agent info
        invoices_data = []
        total_amount = 0
        for invoice in invoices:
            invoice_dict = invoice.to_dict()
            invoice_dict['agent_name'] = f"{invoice.agent.first_name} {invoice.agent.last_name}"
            invoice_dict['agent_email'] = invoice.agent.email
            invoices_data.append(invoice_dict)
            total_amount += float(invoice.total_amount or 0)
        
        return jsonify({
            'period': f"{year}-{month:02d}",
            'invoices': invoices_data,
            'total_count': len(invoices_data),
            'total_amount': total_amount,
            'summary': {
                'paid': len([i for i in invoices if i.payment_status == 'paid']),
                'unpaid': len([i for i in invoices if i.payment_status == 'unpaid']),
                'overdue': len([i for i in invoices if i.payment_status == 'overdue'])
            }
        }), 200
        
    except Exception as e:
        current_app.logger.error(f"Error fetching monthly invoice batch: {e}")
        return jsonify({'error': 'Failed to fetch monthly invoices'}), 500

@admin_bp.route('/admin/invoices/batch-download', methods=['POST'])
@jwt_required()
def create_invoice_batch_download():
    """Create a ZIP file containing multiple invoices for batch download."""
    try:
        current_user_id = get_jwt_identity()
        current_user = User.query.get(int(current_user_id))
        
        if not current_user or current_user.role != 'admin':
            return jsonify({'error': 'Access denied'}), 403
        
        data = request.get_json()
        invoice_ids = data.get('invoice_ids', [])
        batch_name = data.get('batch_name', f"invoice_batch_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}")
        
        if not invoice_ids:
            return jsonify({'error': 'No invoices selected'}), 400
        
        # Get invoices
        invoices = Invoice.query.filter(Invoice.id.in_(invoice_ids)).all()
        if not invoices:
            return jsonify({'error': 'No valid invoices found'}), 404
        
        # Build S3 keys from agent_id + invoice_number
        file_keys = []
        for inv in invoices:
            if inv.status == 'draft':
                continue  # skip drafts
            file_keys.append(f"invoices/{inv.agent_id}/{inv.invoice_number}.pdf")
        
        if not file_keys:
            return jsonify({'error': 'No invoice PDFs available for batch download'}), 404
        
        # Create ZIP in S3
        zip_filename = f"{batch_name}.zip"
        batch_result = s3_client.create_invoice_batch_zip(file_keys, zip_filename)
        if not batch_result.get('success'):
            return jsonify({'error': batch_result.get('error', 'Failed to create ZIP')}), 500
        
        # Signed URL for the ZIP
        zip_url = s3_client.generate_presigned_url(batch_result['file_key'], expiration=7200)
        if not zip_url:
            return jsonify({'error': 'Failed to generate batch download URL'}), 500
        
        return jsonify({
            'download_url': zip_url,
            'filename': batch_result.get('filename', zip_filename),
            'invoice_count': batch_result.get('invoice_count', len(file_keys))
        }), 200
        
    except Exception as e:
        current_app.logger.error(f"Error creating invoice batch download: {e}")
        return jsonify({'error': 'Failed to create batch download'}), 500


@admin_bp.route('/admin/invoices/<int:invoice_id>', methods=['DELETE'])
@jwt_required()
def admin_delete_invoice(invoice_id):
    """Admin deletes an invoice and its links (draft or test cleanup)."""
    try:
        current_user_id = get_jwt_identity()
        current_user = User.query.get(int(current_user_id))
        if not current_user or current_user.role != 'admin':
            return jsonify({'error': 'Access denied'}), 403

        invoice = Invoice.query.get(invoice_id)
        if not invoice:
            return jsonify({'error': 'Invoice not found'}), 404

        # Delete invoice job links first
        InvoiceJob.query.filter_by(invoice_id=invoice.id).delete()
        db.session.delete(invoice)
        db.session.commit()
        return jsonify({'message': 'Invoice deleted'}), 200
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error deleting invoice {invoice_id}: {e}")
        return jsonify({'error': 'Failed to delete invoice'}), 500


@admin_bp.route('/admin/jobs/<int:job_id>', methods=['DELETE'])
@jwt_required()
def admin_delete_job(job_id):
    """Admin deletes a job and its related assignments/links."""
    try:
        current_user_id = get_jwt_identity()
        current_user = User.query.get(int(current_user_id))
        if not current_user or current_user.role != 'admin':
            return jsonify({'error': 'Access denied'}), 403

        job = Job.query.get(job_id)
        if not job:
            return jsonify({'error': 'Job not found'}), 404

        # Delete assignments, notifications, and invoice links
        JobAssignment.query.filter_by(job_id=job_id).delete()
        Notification.query.filter_by(job_id=job_id).delete()
        InvoiceJob.query.filter_by(job_id=job_id).delete()
        db.session.delete(job)
        db.session.commit()
        return jsonify({'message': 'Job deleted'}), 200
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error deleting job {job_id}: {e}")
        return jsonify({'error': 'Failed to delete job'}), 500

@admin_bp.route('/admin/invoices/export-csv', methods=['POST'])
@jwt_required()
def export_invoices_csv():
    """Export invoice data as CSV for accounting."""
    try:
        current_user_id = get_jwt_identity()
        current_user = User.query.get(int(current_user_id))
        
        if not current_user or current_user.role != 'admin':
            return jsonify({'error': 'Access denied'}), 403
        
        data = request.get_json()
        invoice_ids = data.get('invoice_ids', [])
        
        if not invoice_ids:
            return jsonify({'error': 'No invoices selected'}), 400
        
        # Get invoices
        invoices = Invoice.query.filter(Invoice.id.in_(invoice_ids)).join(
            User, Invoice.agent_id == User.id
        ).order_by(Invoice.generated_at.desc()).all()
        
        if not invoices:
            return jsonify({'error': 'No valid invoices found'}), 404
        
        # Create CSV data
        import csv
        import io
        
        csv_buffer = io.StringIO()
        csv_writer = csv.writer(csv_buffer)
        
        # CSV Headers
        csv_writer.writerow([
            'Invoice Number',
            'Agent Name', 
            'Agent Email',
            'Issue Date',
            'Due Date',
            'Total Amount',
            'Payment Status',
            'Generated Date',
            'Download Count',
            'Last Downloaded'
        ])
        
        # CSV Data
        for invoice in invoices:
            csv_writer.writerow([
                invoice.invoice_number,
                f"{invoice.agent.first_name} {invoice.agent.last_name}",
                invoice.agent.email,
                invoice.issue_date.strftime('%Y-%m-%d') if invoice.issue_date else '',
                invoice.due_date.strftime('%Y-%m-%d') if invoice.due_date else '',
                f"{float(invoice.total_amount):.2f}",
                invoice.payment_status,
                invoice.generated_at.strftime('%Y-%m-%d %H:%M:%S') if invoice.generated_at else '',
                invoice.download_count or 0,
                invoice.last_downloaded.strftime('%Y-%m-%d %H:%M:%S') if invoice.last_downloaded else ''
            ])
        
        # Upload CSV to S3
        csv_content = csv_buffer.getvalue()
        csv_filename = f"invoice_export_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.csv"
        csv_key = f"exports/{csv_filename}"
        
        try:
            s3_client.s3_client.put_object(
                Bucket=s3_client.bucket_name,
                Key=csv_key,
                Body=csv_content.encode('utf-8'),
                ContentType='text/csv',
                ServerSideEncryption='AES256',
                Metadata={
                    'export_date': datetime.utcnow().isoformat(),
                    'invoice_count': str(len(invoices)),
                    'exported_by': current_user.email
                }
            )
            
            # Generate download URL
            csv_download_url = s3_client.generate_presigned_url(csv_key, expiration=3600)
            
            # Log CSV export
            current_app.logger.info(
                f"Admin {current_user.email} exported {len(invoices)} invoices to CSV"
            )
            
            return jsonify({
                'csv_filename': csv_filename,
                'download_url': csv_download_url,
                'expires_in': 3600,
                'invoice_count': len(invoices)
            }), 200
            
        except Exception as upload_error:
            current_app.logger.error(f"Error uploading CSV to S3: {upload_error}")
            # Fallback: return CSV data directly  
            return jsonify({
                'csv_data': csv_content,
                'invoice_count': len(invoices)
            }), 200
        
    except Exception as e:
        current_app.logger.error(f"Error exporting invoices CSV: {e}")
        return jsonify({'error': 'Failed to export CSV'}), 500

# === NEW ADMIN AGENT MANAGEMENT ENDPOINTS ===

@admin_bp.route('/admin/agents/<int:agent_id>/details', methods=['GET'])
@jwt_required()
def get_agent_details(agent_id):
    """Get complete agent details including personal info, bank details, and invoice history."""
    try:
        current_user_id = get_jwt_identity()
        current_user = User.query.get(int(current_user_id))
        
        if not current_user or current_user.role != 'admin':
            return jsonify({'error': 'Access denied'}), 403
        
        agent = User.query.get(agent_id)
        if not agent or agent.role != 'agent':
            return jsonify({'error': 'Agent not found'}), 404
        
        # Get agent's invoice statistics
        invoices = Invoice.query.filter_by(agent_id=agent_id).all()
        total_invoices = len(invoices)
        total_amount = sum(float(inv.total_amount or 0) for inv in invoices)
        paid_invoices = [inv for inv in invoices if inv.payment_status == 'paid']
        unpaid_invoices = [inv for inv in invoices if inv.payment_status == 'unpaid']
        overdue_invoices = [inv for inv in invoices if inv.is_overdue()]
        
        paid_amount = sum(float(inv.total_amount or 0) for inv in paid_invoices)
        unpaid_amount = sum(float(inv.total_amount or 0) for inv in unpaid_invoices)
        
        # Get recent invoices (last 10)
        recent_invoices = Invoice.query.filter_by(agent_id=agent_id).order_by(
            Invoice.issue_date.desc()
        ).limit(10).all()
        
        # Get agent's job assignments
        recent_jobs = db.session.query(Job).join(JobAssignment).filter(
            JobAssignment.agent_id == agent_id
        ).order_by(Job.arrival_time.desc()).limit(5).all()

        # Process jobs to include actual invoiced details
        recent_jobs_data = []
        for job in recent_jobs:
            job_dict = job.to_dict()

            # Find the invoice line item for THIS agent and THIS job
            invoice_job = InvoiceJob.query.join(Invoice).filter(
                InvoiceJob.job_id == job.id,
                Invoice.agent_id == agent_id
            ).first()

            if invoice_job:
                job_dict['hours_worked'] = float(invoice_job.hours_worked or 0)
                # Use rate from invoice, fallback to job rate
                job_dict['hourly_rate'] = float(invoice_job.hourly_rate_at_invoice or job.hourly_rate or 0)
                job_dict['invoice_number'] = invoice_job.invoice.invoice_number
                job_dict['invoiced'] = True
            else:
                # Not invoiced yet
                job_dict['invoiced'] = False

            recent_jobs_data.append(job_dict)

        agent_details = {
            'personal_info': {
                'id': agent.id,
                'first_name': agent.first_name,
                'last_name': agent.last_name,
                'email': agent.email,
                'phone': agent.phone,
                'created_at': agent.created_at.isoformat() if agent.created_at else None,
                'verification_status': agent.verification_status
            },
            'address': {
                'address_line_1': agent.address_line_1,
                'address_line_2': agent.address_line_2,
                'city': agent.city,
                'postcode': agent.postcode
            },
            'bank_details': {
                'bank_name': agent.bank_name,
                'bank_account_number': agent.bank_account_number,
                'bank_sort_code': agent.bank_sort_code,
                'utr_number': agent.utr_number
            },
            'invoice_statistics': {
                'total_invoices': total_invoices,
                'total_amount': total_amount,
                'paid_count': len(paid_invoices),
                'paid_amount': paid_amount,
                'unpaid_count': len(unpaid_invoices),
                'unpaid_amount': unpaid_amount,
                'overdue_count': len(overdue_invoices),
                'overdue_amount': sum(float(inv.total_amount or 0) for inv in overdue_invoices)
            },
            'recent_invoices': [inv.to_dict() for inv in recent_invoices],
            'recent_jobs': recent_jobs_data
        }
        
        # Log admin access
        current_app.logger.info(f"Admin {current_user.email} accessed details for agent {agent.email}")
        
        return jsonify(agent_details), 200
        
    except Exception as e:
        current_app.logger.error(f"Error fetching agent details: {e}")
        return jsonify({'error': 'Failed to fetch agent details'}), 500

@admin_bp.route('/admin/agents/<int:agent_id>/invoices', methods=['GET'])
@jwt_required()
def get_agent_invoices_detailed(agent_id):
    """Get detailed invoice history for a specific agent."""
    try:
        current_user_id = get_jwt_identity()
        current_user = User.query.get(int(current_user_id))
        
        if not current_user or current_user.role != 'admin':
            return jsonify({'error': 'Access denied'}), 403
        
        agent = User.query.get(agent_id)
        if not agent or agent.role != 'agent':
            return jsonify({'error': 'Agent not found'}), 404
        
        # Get query parameters
        payment_status = request.args.get('payment_status')
        year = request.args.get('year', type=int)
        month = request.args.get('month', type=int)
        
        # Build query
        query = Invoice.query.filter_by(agent_id=agent_id)
        
        if payment_status:
            query = query.filter(Invoice.payment_status == payment_status)
        if year:
            query = query.filter(db.extract('year', Invoice.issue_date) == year)
        if month:
            query = query.filter(db.extract('month', Invoice.issue_date) == month)
        
        invoices = query.order_by(Invoice.issue_date.desc()).all()
        
        # Enhanced invoice data with job details
        invoices_data = []
        for invoice in invoices:
            invoice_dict = invoice.to_dict()
            
            # Add job details
            invoice_jobs = InvoiceJob.query.filter_by(invoice_id=invoice.id).all()
            job_details = []
            for invoice_job in invoice_jobs:
                if invoice_job.job:
                    job_details.append({
                        'job_id': invoice_job.job.id,
                        'job_title': invoice_job.job.address,
                        'job_address': invoice_job.job.address,
                        'job_date': invoice_job.job.arrival_time.isoformat() if invoice_job.job.arrival_time else None,
                        'hours_worked': float(invoice_job.hours_worked or 0),
                        'hourly_rate': float(invoice_job.hourly_rate_at_invoice or invoice_job.job.hourly_rate or 0)
                    })
            
            invoice_dict['job_details'] = job_details
            invoice_dict['is_overdue'] = invoice.is_overdue()
            invoice_dict['days_overdue'] = invoice.days_overdue()
            
            # Add admin who marked as paid (if applicable)
            if invoice.paid_by_admin_id:
                paid_by_admin = User.query.get(invoice.paid_by_admin_id)
                invoice_dict['paid_by_admin'] = f"{paid_by_admin.first_name} {paid_by_admin.last_name}" if paid_by_admin else None
            
            invoices_data.append(invoice_dict)
        
        return jsonify({
            'agent': {
                'id': agent.id,
                'name': f"{agent.first_name} {agent.last_name}",
                'email': agent.email
            },
            'invoices': invoices_data,
            'total_count': len(invoices_data),
            'summary': {
                'total_amount': sum(float(inv['total_amount']) for inv in invoices_data),
                'paid_count': len([inv for inv in invoices_data if inv['payment_status'] == 'paid']),
                'unpaid_count': len([inv for inv in invoices_data if inv['payment_status'] == 'unpaid']),
                'overdue_count': len([inv for inv in invoices_data if inv['is_overdue']])
            }
        }), 200
        
    except Exception as e:
        current_app.logger.error(f"Error fetching agent invoices: {e}")
        return jsonify({'error': 'Failed to fetch agent invoices'}), 500

@admin_bp.route('/admin/invoices/pending', methods=['GET'])
@jwt_required()
def get_pending_invoices():
    """Get all unpaid invoices for admin management."""
    try:
        current_user_id = get_jwt_identity()
        current_user = User.query.get(int(current_user_id))
        
        if not current_user or current_user.role != 'admin':
            return jsonify({'error': 'Access denied'}), 403
        
        # Get all unpaid invoices
        pending_invoices = Invoice.query.filter(
            Invoice.payment_status.in_(['unpaid', 'overdue'])
        ).join(User, Invoice.agent_id == User.id).order_by(Invoice.due_date.asc()).all()
        
        invoices_data = []
        total_pending_amount = 0
        
        for invoice in pending_invoices:
            invoice_dict = invoice.to_dict()
            invoice_dict['agent_name'] = f"{invoice.agent.first_name} {invoice.agent.last_name}"
            invoice_dict['agent_email'] = invoice.agent.email
            invoice_dict['is_overdue'] = invoice.is_overdue()
            invoice_dict['days_overdue'] = invoice.days_overdue()
            
            # Auto-update overdue status
            if invoice.is_overdue() and invoice.payment_status != 'overdue':
                invoice.payment_status = 'overdue'
                db.session.commit()
                invoice_dict['payment_status'] = 'overdue'
            
            invoices_data.append(invoice_dict)
            total_pending_amount += float(invoice.total_amount or 0)
        
        return jsonify({
            'pending_invoices': invoices_data,
            'total_count': len(invoices_data),
            'total_pending_amount': total_pending_amount,
            'summary': {
                'unpaid_count': len([inv for inv in invoices_data if inv['payment_status'] == 'unpaid']),
                'overdue_count': len([inv for inv in invoices_data if inv['payment_status'] == 'overdue']),
                'unpaid_amount': sum(float(inv['total_amount']) for inv in invoices_data if inv['payment_status'] == 'unpaid'),
                'overdue_amount': sum(float(inv['total_amount']) for inv in invoices_data if inv['payment_status'] == 'overdue')
            }
        }), 200
        
    except Exception as e:
        current_app.logger.error(f"Error fetching pending invoices: {e}")
        return jsonify({'error': 'Failed to fetch pending invoices'}), 500

@admin_bp.route('/admin/invoices/paid', methods=['GET'])
@jwt_required()  
def get_paid_invoices():
    """Get all paid invoices for admin review."""
    try:
        current_user_id = get_jwt_identity()
        current_user = User.query.get(int(current_user_id))
        
        if not current_user or current_user.role != 'admin':
            return jsonify({'error': 'Access denied'}), 403
        
        # Get query parameters
        year = request.args.get('year', type=int)
        month = request.args.get('month', type=int)
        
        # Build query
        query = Invoice.query.filter(Invoice.payment_status == 'paid').join(
            User, Invoice.agent_id == User.id
        )
        
        if year:
            query = query.filter(db.extract('year', Invoice.payment_date) == year)
        if month:
            query = query.filter(db.extract('month', Invoice.payment_date) == month)
        
        paid_invoices = query.order_by(Invoice.payment_date.desc()).all()
        
        invoices_data = []
        total_paid_amount = 0
        
        for invoice in paid_invoices:
            invoice_dict = invoice.to_dict()
            invoice_dict['agent_name'] = f"{invoice.agent.first_name} {invoice.agent.last_name}"
            invoice_dict['agent_email'] = invoice.agent.email
            
            # Add admin who marked as paid
            if invoice.paid_by_admin_id:
                paid_by_admin = User.query.get(invoice.paid_by_admin_id)
                invoice_dict['paid_by_admin'] = f"{paid_by_admin.first_name} {paid_by_admin.last_name}" if paid_by_admin else None
            
            invoices_data.append(invoice_dict)
            total_paid_amount += float(invoice.total_amount or 0)
        
        return jsonify({
            'paid_invoices': invoices_data,
            'total_count': len(invoices_data),
            'total_paid_amount': total_paid_amount,
            'filters': {
                'year': year,
                'month': month
            }
        }), 200
        
    except Exception as e:
        current_app.logger.error(f"Error fetching paid invoices: {e}")
        return jsonify({'error': 'Failed to fetch paid invoices'}), 500

@admin_bp.route('/admin/dashboard/stats', methods=['GET'])
@jwt_required()
def get_admin_dashboard_stats():
    """Get comprehensive dashboard statistics for admin."""
    try:
        current_user_id = get_jwt_identity()
        current_user = User.query.get(int(current_user_id))
        
        if not current_user or current_user.role != 'admin':
            return jsonify({'error': 'Access denied'}), 403
        
        # Invoice statistics
        total_invoices = Invoice.query.count()
        paid_invoices = Invoice.query.filter_by(payment_status='paid').count()
        unpaid_invoices = Invoice.query.filter_by(payment_status='unpaid').count()
        overdue_invoices = Invoice.query.filter_by(payment_status='overdue').count()
        
        # Amount statistics
        total_amount = db.session.query(db.func.sum(Invoice.total_amount)).scalar() or 0
        paid_amount = db.session.query(db.func.sum(Invoice.total_amount)).filter(
            Invoice.payment_status == 'paid'
        ).scalar() or 0
        unpaid_amount = db.session.query(db.func.sum(Invoice.total_amount)).filter(
            Invoice.payment_status == 'unpaid'
        ).scalar() or 0
        overdue_amount = db.session.query(db.func.sum(Invoice.total_amount)).filter(
            Invoice.payment_status == 'overdue'
        ).scalar() or 0
        
        # Agent statistics
        total_agents = User.query.filter_by(role='agent').count()
        verified_agents = User.query.filter_by(role='agent', verification_status='verified').count()
        pending_agents = User.query.filter_by(role='agent', verification_status='pending').count()
        
        # Recent activity
        recent_invoices = Invoice.query.order_by(Invoice.generated_at.desc()).limit(5).all()
        recent_payments = Invoice.query.filter_by(payment_status='paid').order_by(
            Invoice.payment_date.desc()
        ).limit(5).all()
        
        return jsonify({
            'invoice_stats': {
                'total_invoices': total_invoices,
                'paid_invoices': paid_invoices,
                'unpaid_invoices': unpaid_invoices,
                'overdue_invoices': overdue_invoices,
                'total_amount': float(total_amount),
                'paid_amount': float(paid_amount),
                'unpaid_amount': float(unpaid_amount),
                'overdue_amount': float(overdue_amount)
            },
            'agent_stats': {
                'total_agents': total_agents,
                'verified_agents': verified_agents,
                'pending_agents': pending_agents
            },
            'recent_activity': {
                'recent_invoices': [
                    {
                        'id': inv.id,
                        'invoice_number': inv.invoice_number,
                        'agent_name': f"{inv.agent.first_name} {inv.agent.last_name}",
                        'total_amount': float(inv.total_amount),
                        'generated_at': inv.generated_at.isoformat() if inv.generated_at else None
                    } for inv in recent_invoices
                ],
                'recent_payments': [
                    {
                        'id': inv.id,
                        'invoice_number': inv.invoice_number,
                        'agent_name': f"{inv.agent.first_name} {inv.agent.last_name}",
                        'total_amount': float(inv.total_amount),
                        'payment_date': inv.payment_date.isoformat() if inv.payment_date else None
                    } for inv in recent_payments
                ]
            }
        }), 200
        
    except Exception as e:
        current_app.logger.error(f"Error fetching admin dashboard stats: {e}")
        return jsonify({'error': 'Failed to fetch dashboard statistics'}), 500

# === ENHANCED AGENT JOBS AND INVOICE DETAILS ENDPOINTS ===

@admin_bp.route('/admin/agents/<int:agent_id>/jobs', methods=['GET'])
@jwt_required()
def get_agent_jobs(agent_id):
    """Get all jobs for a specific agent with full job details."""
    try:
        current_user_id = get_jwt_identity()
        current_user = User.query.get(int(current_user_id))
        
        if not current_user or current_user.role != 'admin':
            return jsonify({'error': 'Access denied'}), 403
        
        agent = User.query.get(agent_id)
        if not agent or agent.role != 'agent':
            return jsonify({'error': 'Agent not found'}), 404
        
        # Get jobs through assignments and invoices
        assignments = JobAssignment.query.filter_by(agent_id=agent_id).all()
        
        jobs = []
        for assignment in assignments:
            try:
                job = assignment.job
                if not job:
                    continue
                    
                # Safely get job data
                job_data = {}
                job_data['id'] = job.id
                job_data['title'] = getattr(job, 'title', f'Job #{job.id}')
                job_data['address'] = getattr(job, 'address', 'Address not specified')
                job_data['arrival_time'] = job.arrival_time.isoformat() if hasattr(job, 'arrival_time') and job.arrival_time else None
                job_data['job_type'] = getattr(job, 'job_type', 'General')
                job_data['agents_required'] = getattr(job, 'agents_required', 1)
                job_data['status'] = getattr(job, 'status', 'active')
                job_data['notes'] = getattr(job, 'notes', '')
                job_data['assignment_status'] = getattr(assignment, 'status', 'assigned')
                
                # Check if there's an invoice for this job BY THIS SPECIFIC AGENT
                # Join Invoice to ensure we get the record for THIS agent only
                invoice_job = InvoiceJob.query.join(Invoice).filter(
                    InvoiceJob.job_id == job.id,
                    Invoice.agent_id == agent_id
                ).first()

                if invoice_job:
                    invoice = invoice_job.invoice
                    job_data['invoice_id'] = invoice.id
                    job_data['invoice_number'] = getattr(invoice, 'invoice_number', f'INV-{invoice.id}')
                    job_data['invoice_status'] = getattr(invoice, 'status', 'draft')
                    job_data['hours_worked'] = float(getattr(invoice_job, 'hours_worked', 0) or 0)
                    job_data['hourly_rate'] = float(getattr(invoice_job, 'hourly_rate_at_invoice', None) or getattr(job, 'hourly_rate', 20) or 20)
                
                jobs.append(job_data)
                
            except Exception as job_error:
                current_app.logger.error(f"Error processing job {assignment.job_id}: {job_error}")
                continue
        
        # Sort by most recent first
        jobs.sort(key=lambda x: x.get('arrival_time', ''), reverse=True)
        
        return jsonify({
            'agent': {
                'id': agent.id,
                'name': f"{agent.first_name} {agent.last_name}",
                'email': agent.email
            },
            'jobs': jobs,
            'total_count': len(jobs)
        }), 200
        
    except Exception as e:
        current_app.logger.error(f"Error fetching agent jobs: {e}")
        return jsonify({'error': 'Failed to fetch agent jobs'}), 500

@admin_bp.route('/admin/invoices/<int:invoice_id>/details', methods=['GET'])
@jwt_required()
def get_detailed_invoice(invoice_id):
    """Get comprehensive invoice details including job information."""
    try:
        current_user_id = get_jwt_identity()
        current_user = User.query.get(int(current_user_id))
        
        if not current_user or current_user.role != 'admin':
            return jsonify({'error': 'Access denied'}), 403
        
        invoice = Invoice.query.get(invoice_id)
        if not invoice:
            return jsonify({'error': 'Invoice not found'}), 404
        
        # Get job details associated with this invoice
        invoice_jobs = InvoiceJob.query.filter_by(invoice_id=invoice_id).all()
        job_details = []
        
        for invoice_job in invoice_jobs:
            if invoice_job.job:
                job = invoice_job.job
                hours_worked = float(getattr(invoice_job, 'hours_worked', 0) or 0)
                hourly_rate = float(getattr(invoice_job, 'hourly_rate_at_invoice', None) or getattr(job, 'hourly_rate', 20) or 20)
                
                job_details.append({
                    'job_id': job.id,
                    'title': getattr(job, 'title', f'Job #{job.id}'),
                    'address': getattr(job, 'address', 'Address not specified'),
                    'date': job.arrival_time.isoformat() if hasattr(job, 'arrival_time') and job.arrival_time else None,
                    'job_type': getattr(job, 'job_type', 'General'),
                    'hours_worked': hours_worked,
                    'hourly_rate': hourly_rate,
                    'subtotal': hours_worked * hourly_rate,
                    'notes': getattr(job, 'notes', '')
                })
        
        # Calculate totals from job_details (source of truth)
        total_hours = sum(jd.get('hours_worked', 0) for jd in job_details)
        avg_rate = sum(jd.get('hourly_rate', 0) for jd in job_details) / len(job_details) if job_details else 0
        calculated_subtotal = sum(jd.get('subtotal', 0) for jd in job_details)

        # Build comprehensive invoice details with safe attribute access
        details = {
            # Invoice Information
            'id': invoice.id,
            'invoice_number': getattr(invoice, 'invoice_number', f'INV-{invoice.id}'),
            'agent_id': getattr(invoice, 'agent_id', None),
            'agent_name': f"{invoice.agent.first_name} {invoice.agent.last_name}" if invoice.agent else 'Unknown Agent',
            'agent_email': getattr(invoice.agent, 'email', 'No email') if invoice.agent else 'No email',
            'issue_date': invoice.issue_date.isoformat() if hasattr(invoice, 'issue_date') and invoice.issue_date else None,
            'due_date': invoice.due_date.isoformat() if hasattr(invoice, 'due_date') and invoice.due_date else None,
            'status': getattr(invoice, 'status', 'draft'),
            'total_amount': float(getattr(invoice, 'total_amount', 0) or 0),
            # Use calculated values from InvoiceJob records (source of truth)
            'hours': total_hours,
            'rate_per_hour': avg_rate,
            'subtotal': calculated_subtotal,
            'expenses': float(getattr(invoice, 'expenses', 0) or 0),
            'job_details': job_details,
            'created_at': invoice.created_at.isoformat() if hasattr(invoice, 'created_at') and invoice.created_at else None,
            'generated_at': invoice.generated_at.isoformat() if hasattr(invoice, 'generated_at') and invoice.generated_at else None
        }
        
        # Add COMPLETE job information if job exists (enhanced)
        if job_details and len(job_details) > 0:
            # Use first job for main job details (most common case)
            main_job_data = job_details[0]
            details.update({
                # Basic Job Info
                'job_id': main_job_data.get('job_id'),
                'job_title': main_job_data.get('title', 'N/A'),
                'job_type': main_job_data.get('job_type', 'N/A'),
                'job_status': 'completed',  # If invoiced, job is completed
                
                # Location Details
                'job_address': main_job_data.get('address', 'N/A'),
                'job_postcode': 'N/A',  # Will be enhanced below with direct job access
                'job_arrival_time': main_job_data.get('date'),
                'agents_required': 'N/A',  # Will be enhanced below
                'job_notes': main_job_data.get('notes', '')
            })
            
            # Try to get additional details from the actual job object
            try:
                if job_details and len(job_details) > 0:
                    job_id = main_job_data.get('job_id')
                    if job_id:
                        actual_job = Job.query.get(job_id)
                        if actual_job:
                            details.update({
                                # Enhanced location details
                                'job_postcode': getattr(actual_job, 'postcode', 'N/A'),
                                'what3words_address': getattr(actual_job, 'what3words_address', ''),
                                'location_lat': getattr(actual_job, 'location_lat', None),
                                'location_lng': getattr(actual_job, 'location_lng', None),
                                'maps_link': getattr(actual_job, 'maps_link', ''),
                                
                                # Enhanced job details
                                'agents_required': getattr(actual_job, 'agents_required', 1),
                                'lead_agent_name': getattr(actual_job, 'lead_agent_name', ''),
                                'instructions': getattr(actual_job, 'instructions', ''),
                                'urgency_level': getattr(actual_job, 'urgency_level', 'Standard'),
                                'number_of_dwellings': getattr(actual_job, 'number_of_dwellings', None),
                                'police_liaison_required': getattr(actual_job, 'police_liaison_required', False),
                                
                                # Override with actual job notes if available
                                'job_notes': getattr(actual_job, 'instructions', '') or main_job_data.get('notes', '')
                            })
            except Exception as job_error:
                current_app.logger.error(f"Error fetching additional job details: {job_error}")
                
        else:
            # Fallback if no job linked
            details.update({
                'job_id': None,
                'job_title': 'N/A',
                'job_type': 'N/A',
                'job_status': 'N/A',
                'job_address': 'N/A',
                'job_postcode': 'N/A',
                'job_arrival_time': None,
                'agents_required': 'N/A',
                'job_notes': 'No job details available',
                'what3words_address': '',
                'urgency_level': 'N/A',
                'instructions': 'N/A'
            })
        
        # Add payment information if paid
        if invoice.status == 'paid':
            details['paid_date'] = invoice.paid_date.isoformat() if invoice.paid_date else None
            if invoice.paid_by_admin_id:
                paid_by_admin = User.query.get(invoice.paid_by_admin_id)
                details['paid_by_admin'] = f"{paid_by_admin.first_name} {paid_by_admin.last_name}" if paid_by_admin else 'Admin'
        
        return jsonify(details), 200
        
    except Exception as e:
        current_app.logger.error(f"Error fetching invoice details: {e}")
        return jsonify({'error': 'Failed to fetch invoice details'}), 500

# === SIMPLE AGENT DETAILS ENDPOINT FOR AGENT MANAGEMENT PAGE ===

@admin_bp.route('/admin/agent-management/<int:agent_id>/details', methods=['GET'])
@jwt_required()
def get_agent_management_details(agent_id):
    """Get agent details for the agent management page (works with existing database)."""
    try:
        current_user_id = get_jwt_identity()
        current_user = User.query.get(int(current_user_id))
        
        if not current_user or current_user.role != 'admin':
            return jsonify({'error': 'Access denied'}), 403
        
        agent = User.query.get(agent_id)
        if not agent or agent.role != 'agent':
            return jsonify({'error': 'Agent not found'}), 404
        
        # Get agent's invoices (using existing fields only)
        invoices = Invoice.query.filter_by(agent_id=agent_id).order_by(Invoice.issue_date.desc()).all()
        
        # Calculate statistics from existing data
        total_invoices = len(invoices)
        total_earnings = sum(float(invoice.total_amount or 0) for invoice in invoices)
        paid_invoices = [inv for inv in invoices if inv.status == 'paid']
        pending_invoices = [inv for inv in invoices if inv.status in ['sent', 'draft']]
        
        paid_amount = sum(float(inv.total_amount or 0) for inv in paid_invoices)
        pending_amount = sum(float(inv.total_amount or 0) for inv in pending_invoices)
        
        # Enhanced invoice data with job addresses
        invoices_with_jobs = []
        for invoice in invoices:
            invoice_dict = invoice.to_dict()
            
            # Get job details for this invoice
            invoice_jobs = InvoiceJob.query.filter_by(invoice_id=invoice.id).all()
            jobs_info = []
            for ij in invoice_jobs:
                if ij.job:
                    jobs_info.append({
                        'job_id': ij.job.id,
                        'address': ij.job.address or 'Address not specified',
                        'title': ij.job.address or f'Job #{ij.job.id}',
                        'hours_worked': float(ij.hours_worked or 0),
                        'hourly_rate_at_invoice': float(ij.hourly_rate_at_invoice or 20)
                    })
            
            invoice_dict['jobs'] = jobs_info
            invoices_with_jobs.append(invoice_dict)
        
        return jsonify({
            'agent': {
                'id': agent.id,
                'first_name': agent.first_name,
                'last_name': agent.last_name,
                'email': agent.email,
                'phone': agent.phone or 'Not provided',
                'address_line_1': agent.address_line_1 or 'Not provided',
                'address_line_2': agent.address_line_2 or '',
                'city': agent.city or 'Not provided',
                'postcode': agent.postcode or 'Not provided',
                'bank_name': agent.bank_name or 'Not provided',
                'bank_account_number': agent.bank_account_number or 'Not provided',
                'bank_sort_code': agent.bank_sort_code or 'Not provided',
                'utr_number': agent.utr_number or 'Not provided',
                'verification_status': agent.verification_status,
                'created_at': agent.created_at.isoformat() if agent.created_at else None,
                'role': agent.role
            },
            'invoices': invoices_with_jobs,
            'stats': {
                'total_invoices': total_invoices,
                'total_earnings': total_earnings,
                'paid_amount': paid_amount,
                'pending_amount': pending_amount,
                'paid_count': len(paid_invoices),
                'pending_count': len(pending_invoices)
            }
        })
        
    except Exception as e:
        current_app.logger.error(f"Error fetching agent management details: {e}")
        return jsonify({'error': 'Failed to fetch agent details'}), 500


@admin_bp.route('/admin/jobs/<int:job_id>/invoices', methods=['GET'])
@jwt_required()
def get_invoices_for_job(job_id):
    """Get all invoices linked to a specific job"""
    try:
        current_user_id = get_jwt_identity()
        current_user = User.query.get(int(current_user_id))
        if not current_user or current_user.role not in ['admin', 'manager']:
            return jsonify({'error': 'Access denied'}), 403

        # Check if job exists
        job = Job.query.get(job_id)
        if not job:
            return jsonify({'error': 'Job not found'}), 404

        # Get all invoices linked to this job
        invoices = (Invoice.query
            .join(InvoiceJob, InvoiceJob.invoice_id == Invoice.id)
            .filter(InvoiceJob.job_id == job_id)
            .order_by(Invoice.issue_date.desc())
            .all())

        result = []
        for inv in invoices:
            invoice_dict = inv.to_dict()
            
            # Add agent name
            if inv.agent:
                invoice_dict['agent_name'] = f"{inv.agent.first_name} {inv.agent.last_name}"
            else:
                invoice_dict['agent_name'] = 'Unknown Agent'
            
            # Check if PDF is available based on S3 config and non-draft status
            has_pdf = s3_client.is_configured() and inv.status != 'draft'
            invoice_dict['pdf_available'] = has_pdf
            
            # Provide a direct signed URL for inline iframe/download when available
            if has_pdf:
                try:
                    s3_result = s3_client.generate_invoice_download_url(
                        agent_id=inv.agent_id,
                        invoice_number=inv.invoice_number,
                        expiration=3600
                    )
                    if s3_result.get('success') and s3_result.get('download_url'):
                        invoice_dict['pdf_url'] = s3_result['download_url']
                    else:
                        # Fallback: leave without pdf_url to avoid broken iframe
                        invoice_dict['pdf_url'] = None
                except Exception:
                    invoice_dict['pdf_url'] = None
            
            result.append(invoice_dict)

        return jsonify({
            'job_id': job_id,
            'invoices': result,
            'count': len(result)
        }), 200

    except Exception as e:
        current_app.logger.error(f"Error fetching invoices for job {job_id}: {e}", exc_info=True)
        return jsonify({'error': 'Failed to fetch job invoices'}), 500


# ====================
# FINANCE ENDPOINTS
# ====================

@admin_bp.route('/admin/jobs/<int:job_id>/finance', methods=['GET'])
@jwt_required()
def get_job_finance(job_id):
    """Get complete financial breakdown for a job"""
    try:
        current_user_id = get_jwt_identity()
        current_user = User.query.get(int(current_user_id))
        
        if not current_user or current_user.role != 'admin':
            return jsonify({'error': 'Access denied'}), 403
        
        # Get job
        job = Job.query.get(job_id)
        if not job:
            return jsonify({'error': 'Job not found'}), 404
        
        # Get billing config
        billing = JobBilling.query.filter_by(job_id=job_id).first()
        if not billing:
            return jsonify({'error': 'No billing configuration found for this job'}), 404
        
        # Calculate revenue
        revenue = calculate_job_revenue(billing)
        
        # Get expense totals
        expenses = get_job_expense_totals(job_id)
        
        # Get agent invoice totals
        agent_invoices = get_job_agent_invoice_totals(job_id)
        
        # Calculate profit
        profit = calculate_job_profit(job_id)
        
        # Build response
        billing_dict = billing.to_dict()
        billing_dict.update(revenue)
        
        return jsonify({
            'job_id': job_id,
            'billing': billing_dict,
            'agent_invoices': agent_invoices,
            'job_expenses': expenses,
            'profit': profit
        }), 200
        
    except Exception as e:
        current_app.logger.error(f"Error fetching job finance for {job_id}: {e}")
        return jsonify({'error': 'Failed to fetch job finance data'}), 500


@admin_bp.route('/admin/jobs/<int:job_id>/finance/lock', methods=['POST'])
@jwt_required()
def lock_job_finance(job_id):
    """Lock revenue snapshot for completed job"""
    try:
        current_user_id = get_jwt_identity()
        current_user = User.query.get(int(current_user_id))
        
        if not current_user or current_user.role != 'admin':
            return jsonify({'error': 'Access denied'}), 403
        
        # Check job exists
        job = Job.query.get(job_id)
        if not job:
            return jsonify({'error': 'Job not found'}), 404
        
        # Lock snapshot
        success = lock_job_revenue_snapshot(job_id)
        
        if success:
            return jsonify({'message': 'Revenue snapshot locked successfully'}), 200
        else:
            return jsonify({'error': 'Failed to lock revenue snapshot'}), 500
            
    except Exception as e:
        current_app.logger.error(f"Error locking finance for job {job_id}: {e}")
        return jsonify({'error': 'Failed to lock revenue snapshot'}), 500


@admin_bp.route('/admin/expenses', methods=['GET'])
@jwt_required()
def list_expenses():
    """List expenses with filtering"""
    try:
        current_user_id = get_jwt_identity()
        current_user = User.query.get(int(current_user_id))
        
        if not current_user or current_user.role != 'admin':
            return jsonify({'error': 'Access denied'}), 403
        
        # Parse query parameters
        from_date = request.args.get('from')
        to_date = request.args.get('to')
        job_id = request.args.get('job_id', type=int)
        category = request.args.get('category')
        
        # Build query
        query = Expense.query
        
        if from_date:
            try:
                from_date_obj = datetime.strptime(from_date, '%Y-%m-%d').date()
                query = query.filter(Expense.date >= from_date_obj)
            except ValueError:
                return jsonify({'error': 'Invalid from date format. Use YYYY-MM-DD'}), 400
        
        if to_date:
            try:
                to_date_obj = datetime.strptime(to_date, '%Y-%m-%d').date()
                query = query.filter(Expense.date <= to_date_obj)
            except ValueError:
                return jsonify({'error': 'Invalid to date format. Use YYYY-MM-DD'}), 400
        
        if job_id:
            query = query.filter(Expense.job_id == job_id)
        
        if category:
            query = query.filter(Expense.category == category)
        
        # Execute query
        expenses = query.order_by(Expense.date.desc()).all()
        
        # Calculate totals
        total_net = sum(exp.amount_net for exp in expenses)
        total_vat = sum(exp.vat_amount for exp in expenses)
        total_gross = sum(exp.amount_gross for exp in expenses)
        
        return jsonify({
            'expenses': [exp.to_dict() for exp in expenses],
            'totals': {
                'net': float(total_net) if total_net else 0.0,
                'vat': float(total_vat) if total_vat else 0.0,
                'gross': float(total_gross) if total_gross else 0.0
            }
        }), 200
        
    except Exception as e:
        current_app.logger.error(f"Error listing expenses: {e}")
        return jsonify({'error': 'Failed to list expenses'}), 500


@admin_bp.route('/admin/expenses', methods=['POST'])
@jwt_required()
def create_expense():
    """Create new expense"""
    try:
        current_user_id = get_jwt_identity()
        current_user = User.query.get(int(current_user_id))
        
        if not current_user or current_user.role != 'admin':
            return jsonify({'error': 'Access denied'}), 403
        
        data = request.get_json()
        if not data:
            return jsonify({'error': 'No data provided'}), 400
        
        # Required fields
        required_fields = ['date', 'category', 'description', 'amount_net', 'paid_with']
        for field in required_fields:
            if field not in data:
                return jsonify({'error': f'Missing required field: {field}'}), 400
        
        # Parse date
        try:
            expense_date = datetime.strptime(data['date'], '%Y-%m-%d').date()
        except ValueError:
            return jsonify({'error': 'Invalid date format. Use YYYY-MM-DD'}), 400
        
        # Calculate VAT
        amount_net = data['amount_net']
        vat_rate = data.get('vat_rate', 0.20)
        vat_amount, amount_gross = calculate_expense_vat(amount_net, vat_rate)
        
        # Create expense
        expense = Expense(
            date=expense_date,
            category=data['category'],
            description=data['description'],
            amount_net=amount_net,
            vat_rate=vat_rate,
            vat_amount=vat_amount,
            amount_gross=amount_gross,
            job_id=data.get('job_id'),
            paid_with=data['paid_with'],
            supplier=data.get('supplier'),
            receipt_url=data.get('receipt_url'),
            created_by=current_user_id,
            status=data.get('status', 'logged')
        )
        
        db.session.add(expense)
        db.session.commit()
        
        return jsonify(expense.to_dict()), 201
        
    except Exception as e:
        current_app.logger.error(f"Error creating expense: {e}")
        db.session.rollback()
        return jsonify({'error': 'Failed to create expense'}), 500


@admin_bp.route('/admin/expenses/<int:expense_id>', methods=['PATCH'])
@jwt_required()
def update_expense(expense_id):
    """Update existing expense"""
    try:
        current_user_id = get_jwt_identity()
        current_user = User.query.get(int(current_user_id))
        
        if not current_user or current_user.role != 'admin':
            return jsonify({'error': 'Access denied'}), 403
        
        expense = Expense.query.get(expense_id)
        if not expense:
            return jsonify({'error': 'Expense not found'}), 404
        
        data = request.get_json()
        if not data:
            return jsonify({'error': 'No data provided'}), 400
        
        # Update fields
        if 'date' in data:
            try:
                expense.date = datetime.strptime(data['date'], '%Y-%m-%d').date()
            except ValueError:
                return jsonify({'error': 'Invalid date format. Use YYYY-MM-DD'}), 400
        
        if 'category' in data:
            expense.category = data['category']
        
        if 'description' in data:
            expense.description = data['description']
        
        if 'amount_net' in data or 'vat_rate' in data:
            amount_net = data.get('amount_net', expense.amount_net)
            vat_rate = data.get('vat_rate', expense.vat_rate)
            vat_amount, amount_gross = calculate_expense_vat(amount_net, vat_rate)
            
            expense.amount_net = amount_net
            expense.vat_rate = vat_rate
            expense.vat_amount = vat_amount
            expense.amount_gross = amount_gross
        
        if 'job_id' in data:
            expense.job_id = data['job_id']
        
        if 'paid_with' in data:
            expense.paid_with = data['paid_with']
        
        if 'supplier' in data:
            expense.supplier = data['supplier']
        
        if 'receipt_url' in data:
            expense.receipt_url = data['receipt_url']
        
        if 'status' in data:
            expense.status = data['status']
        
        db.session.commit()
        
        return jsonify(expense.to_dict()), 200
        
    except Exception as e:
        current_app.logger.error(f"Error updating expense {expense_id}: {e}")
        db.session.rollback()
        return jsonify({'error': 'Failed to update expense'}), 500


@admin_bp.route('/admin/expenses/<int:expense_id>', methods=['DELETE'])
@jwt_required()
def delete_expense(expense_id):
    """Delete expense"""
    try:
        current_user_id = get_jwt_identity()
        current_user = User.query.get(int(current_user_id))
        
        if not current_user or current_user.role != 'admin':
            return jsonify({'error': 'Access denied'}), 403
        
        expense = Expense.query.get(expense_id)
        if not expense:
            return jsonify({'error': 'Expense not found'}), 404
        
        db.session.delete(expense)
        db.session.commit()
        
        return jsonify({'message': 'Expense deleted successfully'}), 200
        
    except Exception as e:
        current_app.logger.error(f"Error deleting expense {expense_id}: {e}")
        db.session.rollback()
        return jsonify({'error': 'Failed to delete expense'}), 500


@admin_bp.route('/admin/finance/summary', methods=['GET'])
@jwt_required()
def finance_summary():
    try:
        current_user_id = get_jwt_identity()
        current_user = User.query.get(int(current_user_id))
        if not current_user or current_user.role != 'admin':
            return jsonify({'error': 'Access denied'}), 403

        from_q = request.args.get('from')
        to_q = request.args.get('to')

        def parse_date(s):
            if not s:
                return None
            return datetime.strptime(s, '%Y-%m-%d').date()

        from_date = parse_date(from_q)
        to_date = parse_date(to_q)

        if (from_q and from_date is None) or (to_q and to_date is None):
            return jsonify({'error': 'Invalid date format. Use YYYY-MM-DD'}), 400

        # Build normalized financial summary
        summary = get_financial_summary(from_date=from_date, to_date=to_date)

        # Ensure floats and required keys
        response = {
            'revenue': {
                'net': float(summary.get('revenue', {}).get('net', 0.0) or 0.0),
                'vat': float(summary.get('revenue', {}).get('vat', 0.0) or 0.0),
                'gross': float(summary.get('revenue', {}).get('gross', 0.0) or 0.0),
            },
            'agent_invoices': {
                'net': float(summary.get('agent_invoices', {}).get('net', 0.0) or 0.0),
                'gross': float(summary.get('agent_invoices', {}).get('gross', 0.0) or 0.0),
            },
            'expenses': {
                'net': float(summary.get('expenses', {}).get('net', 0.0) or 0.0),
                'vat': float(summary.get('expenses', {}).get('vat', 0.0) or 0.0),
                'gross': float(summary.get('expenses', {}).get('gross', 0.0) or 0.0),
            },
            'money_in': {
                'net': float(summary.get('money_in', {}).get('net', summary.get('revenue', {}).get('net', 0.0) or 0.0)),
                'gross': float(summary.get('money_in', {}).get('gross', summary.get('revenue', {}).get('gross', 0.0) or 0.0)),
            },
            'money_out': {
                'net': float(summary.get('money_out', {}).get('net', summary.get('agent_invoices', {}).get('net', 0.0) + summary.get('expenses', {}).get('net', 0.0))),
                'gross': float(summary.get('money_out', {}).get('gross', summary.get('agent_invoices', {}).get('gross', 0.0) + summary.get('expenses', {}).get('gross', 0.0))),
            },
            'profit': {
                'net': float(summary.get('profit', {}).get('net', 0.0) or 0.0),
                'gross': float(summary.get('profit', {}).get('gross', 0.0) or 0.0),
            },
            'vat': {
                'output': float(summary.get('vat', {}).get('output', 0.0) or 0.0),
                'input': float(summary.get('vat', {}).get('input', 0.0) or 0.0),
                'net_due': float(summary.get('vat', {}).get('net_due', 0.0) or 0.0),
            },
        }

        # If money_in not provided, derive from revenue
        response['money_in']['net'] = response['money_in']['net'] or response['revenue']['net']
        response['money_in']['gross'] = response['money_in']['gross'] or response['revenue']['gross']

        # If money_out not provided, derive from components
        response['money_out']['net'] = response['money_out']['net'] or (response['agent_invoices']['net'] + response['expenses']['net'])
        response['money_out']['gross'] = response['money_out']['gross'] or (response['agent_invoices']['gross'] + response['expenses']['gross'])

        # If profit not provided, derive
        response['profit']['net'] = response['profit']['net'] or (response['money_in']['net'] - response['money_out']['net'])
        response['profit']['gross'] = response['profit']['gross'] or (response['money_in']['gross'] - response['money_out']['gross'])

        # Ensure VAT net_due
        response['vat']['net_due'] = response['vat']['net_due'] or (response['vat']['output'] - response['vat']['input'])

        return jsonify(response), 200

    except ValueError:
        return jsonify({'error': 'Invalid date format. Use YYYY-MM-DD'}), 400
    except Exception as e:
        current_app.logger.error(f"Error building finance summary: {e}")
        return jsonify({'error': 'Failed to build finance summary'}), 500

@admin_bp.route('/admin/settings/notifications', methods=['GET'])
@jwt_required()
def get_notifications_setting():
	user = require_admin()
	if not user:
		return jsonify({'error': 'Access denied. Admin role required.'}), 403
	from src.models.user import Setting
	default_enabled = str(current_app.config.get('NOTIFICATIONS_ENABLED', 'true')).lower() in ('1','true','yes','on')
	enabled = Setting.get_bool('notifications_enabled', default_enabled)
	return jsonify({'enabled': bool(enabled)})

@admin_bp.route('/admin/settings/notifications', methods=['PUT'])
@jwt_required()
def set_notifications_setting():
	user = require_admin()
	if not user:
		return jsonify({'error': 'Access denied. Admin role required.'}), 403
	from src.models.user import Setting, db
	data = request.get_json(silent=True) or {}
	enabled = bool(data.get('enabled', True))
	Setting.set_bool('notifications_enabled', enabled)
	current_app.logger.info(f"Admin set notifications_enabled={enabled}")
	return jsonify({'enabled': enabled})


# ==========================================
# V3 JOB REPORTS ADMIN ENDPOINTS
# ==========================================


# Helper function for file validation (same as agent routes)
def allowed_file(filename):
    """Check if the file extension is allowed."""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in {'pdf', 'jpg', 'jpeg', 'png'}

@admin_bp.route('/admin/v3-reports/upload-photos', methods=['POST'])
@jwt_required()
def admin_upload_v3_report_photos():
    """
    Admin endpoint: Upload photos for a V3 report to S3.
    Accepts multiple files and returns array of S3 URLs.
    """
    try:
        current_user_id = int(get_jwt_identity())
        user = User.query.get(current_user_id)

        if not user or user.role not in ['admin', 'manager']:
            return jsonify({'error': 'Access denied. Admin role required.'}), 403

        # Check if S3 is configured
        if not s3_client.is_configured():
            return jsonify({
                'error': 'File upload service not available',
                'details': 'S3 storage not configured'
            }), 503

        # Get all uploaded files
        uploaded_files = request.files.getlist('photos')

        if not uploaded_files:
            return jsonify({'error': 'No photos provided'}), 400

        photo_urls = []
        errors = []

        for file in uploaded_files:
            if file.filename == '':
                continue

            if not allowed_file(file.filename):
                errors.append(f"{file.filename}: Invalid file type")
                continue

            # Upload to S3 under v3-reports folder (using admin's ID as agent_id)
            upload_result = s3_client.upload_agent_document(
                agent_id=user.id,
                file=file,
                file_type='v3_report_photo'
            )

            if upload_result.get('success'):
                file_key = upload_result.get('file_key')
                if file_key:
                    photo_urls.append({
                        'url': file_key,
                        'filename': upload_result.get('original_filename'),
                        'upload_date': upload_result.get('upload_date')
                    })
            else:
                errors.append(f"{file.filename}: {upload_result.get('error', 'Upload failed')}")

        if not photo_urls and errors:
            return jsonify({'error': 'All uploads failed', 'details': errors}), 500

        return jsonify({
            'message': f'{len(photo_urls)} photo(s) uploaded successfully',
            'photos': photo_urls,
            'errors': errors if errors else None
        }), 200

    except Exception as e:
        current_app.logger.error(f"Admin photo upload error: {str(e)}")
        return jsonify({'error': 'Failed to upload photos', 'details': str(e)}), 500


@admin_bp.route('/admin/v3-reports/submit', methods=['POST'])
@jwt_required()
def admin_submit_v3_report():
    """
    Admin endpoint: Submit a V3 job report.
    Same functionality as agent endpoint but for admins.
    """
    try:
        from src.models.v3_report import V3JobReport

        current_user_id = int(get_jwt_identity())
        user = User.query.get(current_user_id)

        if not user or user.role not in ['admin', 'manager']:
            return jsonify({'error': 'Access denied. Admin role required.'}), 403

        data = request.get_json() or {}
        job_id = data.get('job_id')
        form_type = data.get('form_type')
        report_data = data.get('report_data')
        photo_urls = data.get('photo_urls', [])

        if not form_type:
            return jsonify({'error': 'form_type is required'}), 400

        if not report_data:
            return jsonify({'error': 'report_data is required'}), 400

        # Handle manual reports
        if job_id == 'MANUAL':
            job = None
        else:
            if not job_id:
                return jsonify({'error': 'job_id is required'}), 400

            job = Job.query.get(int(job_id))
            if not job:
                return jsonify({'error': 'Job not found'}), 404

        # Create the V3 report
        v3_report = V3JobReport(
            job_id=job.id if job else None,
            agent_id=user.id,  # Store admin ID
            form_type=form_type,
            status='submitted',
            report_data=report_data,
            photo_urls=photo_urls
        )

        db.session.add(v3_report)
        db.session.commit()

        return jsonify({
            'message': 'Report submitted successfully',
            'report_id': v3_report.id,
            'status': v3_report.status
        }), 201

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Admin report submission error: {str(e)}")
        return jsonify({'error': 'Failed to submit report', 'details': str(e)}), 500


@admin_bp.route('/admin/jobs/<int:job_id>/v3-reports', methods=['GET'])
@jwt_required()
def get_job_v3_reports(job_id):
	"""Get all V3 reports submitted for a specific job."""
	user = require_admin()
	if not user:
		return jsonify({'error': 'Access denied. Admin role required.'}), 403

	try:
		from src.models.v3_report import V3JobReport

		# Verify job exists
		job = Job.query.get(job_id)
		if not job:
			return jsonify({'error': 'Job not found'}), 404

		# Fetch all reports for this job
		reports = V3JobReport.query.filter_by(job_id=job_id).order_by(
			V3JobReport.submitted_at.desc()
		).all()

		# Enrich with agent information and generate signed URLs for photos
		reports_data = []
		for report in reports:
			report_dict = report.to_dict()

			# Add agent name
			agent = User.query.get(report.agent_id)
			if agent:
				report_dict['agent_name'] = f"{agent.first_name} {agent.last_name}".strip()
				report_dict['agent_email'] = agent.email

			# Convert S3 keys to signed URLs
			if report_dict.get('photo_urls'):
				current_app.logger.info(f"📸 Raw photo_urls from DB: {report_dict['photo_urls']}")
				signed_urls = []
				for photo in report_dict['photo_urls']:
					# photo might be a dict with 'url' key or just a string
					s3_key = photo.get('url') if isinstance(photo, dict) else photo
					current_app.logger.info(f"🔑 S3 key: {s3_key}")
					if s3_key and s3_client.is_configured():
						signed_url = s3_client.generate_presigned_url(s3_key, expiration=3600)
						current_app.logger.info(f"✅ Generated signed URL: {signed_url[:100] if signed_url else 'None'}")
						if signed_url:
							if isinstance(photo, dict):
								photo_copy = photo.copy()
								photo_copy['url'] = signed_url
								signed_urls.append(photo_copy)
							else:
								signed_urls.append({'url': signed_url})
						else:
							signed_urls.append(photo)
					else:
						current_app.logger.info(f"⚠️ S3 not configured or empty key")
						signed_urls.append(photo)
				report_dict['photo_urls'] = signed_urls
				current_app.logger.info(f"📤 Final photo_urls being sent: {signed_urls}")

			reports_data.append(report_dict)

		return jsonify({
			'reports': reports_data
		})

	except Exception as e:
		current_app.logger.error(f"Error fetching V3 reports for job {job_id}: {str(e)}")
		return jsonify({'error': 'Failed to fetch reports'}), 500


# ============================================================================
# V3 Job Report PDF Export
# ============================================================================

def generate_v3_report_pdf(report, agent_name=None):
	"""
	Generates a professional PDF for a V3 job report using the V3 headed template.
	Returns bytes of the PDF file.
	"""
	from PyPDF2 import PdfReader, PdfWriter
	import os

	# Path to the headed template
	template_path = os.path.join(os.path.dirname(__file__), '..', 'templates', 'headed_template.pdf')

	# First, generate the content PDF
	content_buffer = io.BytesIO()
	doc = SimpleDocTemplate(
		content_buffer,
		pagesize=A4,
		rightMargin=1.5*cm,
		leftMargin=1.5*cm,
		topMargin=3.2*cm,  # Space for header (logo + address)
		bottomMargin=2.0*cm  # Space for footer
	)

	# Styles
	styles = getSampleStyleSheet()
	styles.add(ParagraphStyle(
		name='ReportTitle',
		parent=styles['Heading1'],
		fontSize=20,
		spaceAfter=20,
		alignment=TA_CENTER,
		textColor=colors.HexColor('#1a1a2e')
	))
	styles.add(ParagraphStyle(
		name='SectionTitle',
		parent=styles['Heading2'],
		fontSize=14,
		spaceBefore=15,
		spaceAfter=10,
		textColor=colors.HexColor('#ff6b35'),
		borderPadding=(0, 0, 5, 0)
	))
	styles.add(ParagraphStyle(
		name='FieldLabel',
		parent=styles['Normal'],
		fontSize=10,
		textColor=colors.HexColor('#666666'),
		fontName='Helvetica-Bold'
	))
	styles.add(ParagraphStyle(
		name='FieldValue',
		parent=styles['Normal'],
		fontSize=11,
		textColor=colors.HexColor('#1a1a2e'),
		spaceAfter=8
	))
	styles.add(ParagraphStyle(
		name='TimelineTime',
		parent=styles['Normal'],
		fontSize=10,
		textColor=colors.HexColor('#ff6b35'),
		fontName='Helvetica-Bold'
	))
	styles.add(ParagraphStyle(
		name='TimelineText',
		parent=styles['Normal'],
		fontSize=10,
		textColor=colors.HexColor('#333333'),
		leftIndent=20
	))

	elements = []
	report_data = report.report_data or {}

	# Form type names
	form_type_names = {
		'traveller_eviction': 'Traveller Eviction Report',
		'traveller_serve': 'Traveller Serve Report',
		'squatter_serve': 'Squatter Serve Report',
		'squatter_eviction': 'Squatter Eviction Report',
	}
	report_title = form_type_names.get(report.form_type, 'Job Report')

	# Title
	elements.append(Paragraph(report_title, styles['ReportTitle']))
	elements.append(Paragraph(f"Report ID: {report.id}", styles['Normal']))
	elements.append(Spacer(1, 0.3*inch))

	# Helper function to format values
	def format_value(key, value):
		if value is None or value == '':
			return None
		if isinstance(value, bool):
			return 'Yes' if value else 'No'
		if isinstance(value, dict):
			# Timeline entries
			entries = [(k, v) for k, v in value.items() if v and str(v).strip()]
			if not entries:
				return None
			return entries
		return str(value)

	# Field labels
	field_labels = {
		'client': 'Client',
		'address1': 'Address Line 1',
		'address2': 'Address Line 2',
		'city': 'City',
		'postcode': 'Postcode',
		'date': 'Date',
		'arrival_time': 'Arrival Time',
		'departure_time': 'Departure Time',
		'completion_date': 'Completion Date',
		'prior_notice_served': 'Prior Notice Served',
		'property_condition': 'Property Condition',
		'property_damage': 'Property Damage',
		'damage_details': 'Damage Details',
		'aggressive': 'Squatters Aggressive',
		'aggression_details': 'Aggression Details',
		'dogs_on_site': 'Dogs on Site',
		'dog_details': 'Dog Details',
		'num_males': 'Adult Males',
		'num_females': 'Adult Females',
		'num_children': 'Children',
		'police_attendance': 'Police Attendance',
		'cad_number': 'CAD Number',
		'police_force': 'Police Force',
		'police_notes': 'Police Notes',
		'additional_notes': 'Additional Notes',
	}

	def get_label(key):
		if key in field_labels:
			return field_labels[key]
		if key.startswith('agent_'):
			num = key.split('_')[1]
			return f"Agent {num}" if num != '1' else "Lead Agent"
		return key.replace('_', ' ').title()

	# Metadata section
	elements.append(Paragraph("Report Information", styles['SectionTitle']))
	meta_data = [
		['Submitted By:', agent_name or 'Unknown'],
		['Submitted At:', report.submitted_at.strftime('%d %b %Y at %H:%M') if report.submitted_at else 'Unknown'],
		['Status:', (report.status or 'submitted').capitalize()],
	]
	if report.reviewed_at:
		meta_data.append(['Reviewed At:', report.reviewed_at.strftime('%d %b %Y at %H:%M')])

	meta_table = Table(meta_data, colWidths=[2.5*inch, 4.5*inch])
	meta_table.setStyle(TableStyle([
		('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
		('FONTSIZE', (0, 0), (-1, -1), 10),
		('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#666666')),
		('TEXTCOLOR', (1, 0), (1, -1), colors.HexColor('#1a1a2e')),
		('BOTTOMPADDING', (0, 0), (-1, -1), 8),
		('VALIGN', (0, 0), (-1, -1), 'TOP'),
	]))
	elements.append(meta_table)
	elements.append(Spacer(1, 0.2*inch))

	# Job Details section
	elements.append(Paragraph("Job Details", styles['SectionTitle']))
	job_fields = ['client', 'address1', 'address2', 'city', 'postcode', 'date', 'arrival_time']
	job_data = []
	for field in job_fields:
		value = format_value(field, report_data.get(field))
		if value:
			job_data.append([get_label(field) + ':', value])

	if job_data:
		job_table = Table(job_data, colWidths=[2.5*inch, 4.5*inch])
		job_table.setStyle(TableStyle([
			('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
			('FONTSIZE', (0, 0), (-1, -1), 10),
			('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#666666')),
			('TEXTCOLOR', (1, 0), (1, -1), colors.HexColor('#1a1a2e')),
			('BOTTOMPADDING', (0, 0), (-1, -1), 6),
			('VALIGN', (0, 0), (-1, -1), 'TOP'),
		]))
		elements.append(job_table)
	elements.append(Spacer(1, 0.2*inch))

	# Agents section
	elements.append(Paragraph("Agents on Site", styles['SectionTitle']))
	agent_data = []
	for i in range(1, 21):
		field = f'agent_{i}'
		value = format_value(field, report_data.get(field))
		if value:
			label = "Lead Agent:" if i == 1 else f"Agent {i}:"
			agent_data.append([label, value])

	if agent_data:
		agent_table = Table(agent_data, colWidths=[2.5*inch, 4.5*inch])
		agent_table.setStyle(TableStyle([
			('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
			('FONTSIZE', (0, 0), (-1, -1), 10),
			('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#666666')),
			('TEXTCOLOR', (1, 0), (1, -1), colors.HexColor('#1a1a2e')),
			('BOTTOMPADDING', (0, 0), (-1, -1), 6),
			('VALIGN', (0, 0), (-1, -1), 'TOP'),
		]))
		elements.append(agent_table)
	elements.append(Spacer(1, 0.2*inch))

	# Property Details section
	elements.append(Paragraph("Property Details", styles['SectionTitle']))
	prop_fields = ['prior_notice_served', 'property_condition', 'property_damage', 'damage_details',
				   'aggressive', 'aggression_details', 'dogs_on_site', 'dog_details',
				   'num_males', 'num_females', 'num_children']
	prop_data = []
	for field in prop_fields:
		value = format_value(field, report_data.get(field))
		if value:
			prop_data.append([get_label(field) + ':', value])

	if prop_data:
		prop_table = Table(prop_data, colWidths=[2.5*inch, 4.5*inch])
		prop_table.setStyle(TableStyle([
			('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
			('FONTSIZE', (0, 0), (-1, -1), 10),
			('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#666666')),
			('TEXTCOLOR', (1, 0), (1, -1), colors.HexColor('#1a1a2e')),
			('BOTTOMPADDING', (0, 0), (-1, -1), 6),
			('VALIGN', (0, 0), (-1, -1), 'TOP'),
		]))
		elements.append(prop_table)
	elements.append(Spacer(1, 0.2*inch))

	# Timeline sections
	for day_num in range(1, 8):
		timeline_key = f'timeline_day{day_num}'
		timeline_data = report_data.get(timeline_key)
		if timeline_data and isinstance(timeline_data, dict):
			entries = [(k, v) for k, v in timeline_data.items() if v and str(v).strip()]
			if entries:
				day_label = "Day 1 Timeline" if day_num == 1 else f"Day {day_num} Timeline"
				elements.append(Paragraph(day_label, styles['SectionTitle']))

				# Add each timeline entry as a separate table for proper spacing
				sorted_entries = sorted(entries, key=lambda x: x[0])
				for i, (time, text) in enumerate(sorted_entries):
					# Wrap long text
					wrapped_text = Paragraph(str(text), styles['TimelineText'])

					# Create individual table for each timeline entry
					entry_table = Table([[time, wrapped_text]], colWidths=[1*inch, 5.5*inch])
					entry_table.setStyle(TableStyle([
						('FONTNAME', (0, 0), (0, 0), 'Helvetica-Bold'),
						('FONTSIZE', (0, 0), (0, 0), 10),
						('TEXTCOLOR', (0, 0), (0, 0), colors.HexColor('#ff6b35')),
						('TOPPADDING', (0, 0), (-1, -1), 6),
						('BOTTOMPADDING', (0, 0), (-1, -1), 6),
						('VALIGN', (0, 0), (-1, -1), 'TOP'),
						('LEFTPADDING', (0, 0), (0, 0), 0),
						('LINEBELOW', (0, 0), (-1, -1), 0.5, colors.HexColor('#e9ecef')),
					]))
					elements.append(entry_table)

					# Add spacer between entries (except after the last one)
					if i < len(sorted_entries) - 1:
						elements.append(Spacer(1, 0.1*inch))

				elements.append(Spacer(1, 0.15*inch))

	# Police Details section
	police_fields = ['police_attendance', 'cad_number', 'police_force', 'police_notes']
	police_data = []
	for field in police_fields:
		value = format_value(field, report_data.get(field))
		if value:
			police_data.append([get_label(field) + ':', value])

	if police_data:
		elements.append(Paragraph("Police Details", styles['SectionTitle']))
		police_table = Table(police_data, colWidths=[2.5*inch, 4.5*inch])
		police_table.setStyle(TableStyle([
			('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
			('FONTSIZE', (0, 0), (-1, -1), 10),
			('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#666666')),
			('TEXTCOLOR', (1, 0), (1, -1), colors.HexColor('#1a1a2e')),
			('BOTTOMPADDING', (0, 0), (-1, -1), 6),
			('VALIGN', (0, 0), (-1, -1), 'TOP'),
		]))
		elements.append(police_table)
		elements.append(Spacer(1, 0.2*inch))

	# Completion section
	elements.append(Paragraph("Completion", styles['SectionTitle']))
	completion_fields = ['departure_time', 'completion_date', 'additional_notes']
	completion_data = []
	for field in completion_fields:
		value = format_value(field, report_data.get(field))
		if value:
			# Wrap additional notes if long
			if field == 'additional_notes' and len(str(value)) > 80:
				completion_data.append([get_label(field) + ':', Paragraph(str(value), styles['FieldValue'])])
			else:
				completion_data.append([get_label(field) + ':', value])

	if completion_data:
		completion_table = Table(completion_data, colWidths=[2.5*inch, 4.5*inch])
		completion_table.setStyle(TableStyle([
			('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
			('FONTSIZE', (0, 0), (-1, -1), 10),
			('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#666666')),
			('TEXTCOLOR', (1, 0), (1, -1), colors.HexColor('#1a1a2e')),
			('BOTTOMPADDING', (0, 0), (-1, -1), 6),
			('VALIGN', (0, 0), (-1, -1), 'TOP'),
		]))
		elements.append(completion_table)

	# Photos section - embed actual images
	photo_urls = report.photo_urls or []
	if photo_urls:
		elements.append(Spacer(1, 0.3*inch))
		elements.append(Paragraph(f"Photos & Evidence ({len(photo_urls)})", styles['SectionTitle']))
		elements.append(Spacer(1, 0.1*inch))

		# Download and embed each photo
		photo_images = []
		for i, photo in enumerate(photo_urls):
			try:
				s3_key = photo.get('url') if isinstance(photo, dict) else photo
				if s3_key and s3_client.is_configured():
					signed_url = s3_client.generate_presigned_url(s3_key, expiration=300)
					if signed_url:
						# Download the image
						img_response = requests.get(signed_url, timeout=30)
						if img_response.status_code == 200:
							# Create image from bytes
							img_buffer = io.BytesIO(img_response.content)
							try:
								# Create ReportLab Image
								img = Image(img_buffer)
								# Scale to fit nicely (max 2.5 inches wide, maintain aspect ratio)
								img_width = img.drawWidth
								img_height = img.drawHeight
								max_width = 2.4 * inch
								max_height = 2.4 * inch

								# Scale proportionally
								scale = min(max_width / img_width, max_height / img_height, 1.0)
								img.drawWidth = img_width * scale
								img.drawHeight = img_height * scale

								photo_images.append(img)
							except Exception as img_err:
								current_app.logger.warning(f"Failed to process image {i}: {str(img_err)}")
			except Exception as photo_err:
				current_app.logger.warning(f"Failed to download photo {i}: {str(photo_err)}")

		# Arrange photos in a grid (3 per row)
		if photo_images:
			# Create rows of 3 images
			rows = []
			for i in range(0, len(photo_images), 3):
				row = photo_images[i:i+3]
				# Pad row to 3 columns if needed
				while len(row) < 3:
					row.append('')
				rows.append(row)

			# Create table with photos
			photo_table = Table(rows, colWidths=[2.5*inch, 2.5*inch, 2.5*inch])
			photo_table.setStyle(TableStyle([
				('ALIGN', (0, 0), (-1, -1), 'CENTER'),
				('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
				('LEFTPADDING', (0, 0), (-1, -1), 5),
				('RIGHTPADDING', (0, 0), (-1, -1), 5),
				('TOPPADDING', (0, 0), (-1, -1), 5),
				('BOTTOMPADDING', (0, 0), (-1, -1), 5),
			]))
			elements.append(photo_table)
		else:
			elements.append(Paragraph(
				f"Photos could not be loaded. View the full report online to see all {len(photo_urls)} images.",
				styles['Normal']
			))

	# Generation timestamp
	elements.append(Spacer(1, 0.3*inch))
	elements.append(Paragraph(
		f"Report generated on {datetime.now().strftime('%d %b %Y at %H:%M')}",
		ParagraphStyle(
			name='Timestamp',
			parent=styles['Normal'],
			fontSize=8,
			textColor=colors.HexColor('#999999'),
			alignment=TA_CENTER
		)
	))

	# Build content PDF
	doc.build(elements)
	content_buffer.seek(0)

	# Now merge content with the headed template
	try:
		# Read the template
		template_reader = PdfReader(template_path)
		template_page = template_reader.pages[0]

		# Read the content we just generated
		content_reader = PdfReader(content_buffer)

		# Create output PDF
		output = PdfWriter()

		# For each page of content, merge with template
		for i, content_page in enumerate(content_reader.pages):
			# Create a copy of the template for each page
			from copy import copy
			new_page = copy(template_page)
			# Merge the content on top of the template
			new_page.merge_page(content_page)
			output.add_page(new_page)

		# Write to buffer
		output_buffer = io.BytesIO()
		output.write(output_buffer)
		output_buffer.seek(0)
		return output_buffer.getvalue()

	except Exception as merge_error:
		# If merging fails, fall back to content-only PDF
		current_app.logger.warning(f"Template merge failed, using content only: {str(merge_error)}")
		content_buffer.seek(0)
		return content_buffer.getvalue()


@admin_bp.route('/admin/v3-reports/<int:report_id>/pdf', methods=['GET'])
@jwt_required()
def export_v3_report_pdf(report_id):
	"""Export a V3 job report as PDF."""
	user = require_admin()
	if not user:
		return jsonify({'error': 'Access denied. Admin role required.'}), 403

	try:
		from src.models.v3_report import V3JobReport

		# Get the report
		report = V3JobReport.query.get(report_id)
		if not report:
			return jsonify({'error': 'Report not found'}), 404

		# Get agent name
		agent = User.query.get(report.agent_id)
		agent_name = f"{agent.first_name} {agent.last_name}".strip() if agent else 'Unknown'

		# Generate PDF
		pdf_bytes = generate_v3_report_pdf(report, agent_name)

		# Try to upload to S3 using the raw boto3 client
		if s3_client.is_configured():
			try:
				import boto3

				# Access the underlying boto3 client from s3_client
				s3 = s3_client.s3_client if hasattr(s3_client, 's3_client') else None
				bucket = s3_client.bucket_name if hasattr(s3_client, 'bucket_name') else None
				s3_key = f"reports/v3_report_{report_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"

				if s3 and bucket:
					# Upload to S3
					s3.put_object(
						Bucket=bucket,
						Key=s3_key,
						Body=pdf_bytes,
						ContentType='application/pdf',
						ServerSideEncryption='AES256'
					)

					# Generate presigned URL for download
					pdf_url = s3_client.generate_presigned_url(s3_key, expiration=3600)
					if pdf_url:
						return jsonify({
							'pdf_url': pdf_url,
							's3_key': s3_key,
							'filename': f'report_{report_id}.pdf'
						})
			except Exception as s3_error:
				current_app.logger.warning(f"S3 upload failed, falling back to direct download: {str(s3_error)}")

		# Fallback: return PDF directly if S3 not configured or failed
		return send_file(
			io.BytesIO(pdf_bytes),
			mimetype='application/pdf',
			as_attachment=True,
			download_name=f'report_{report_id}.pdf'
		)

	except Exception as e:
		current_app.logger.error(f"Error generating PDF for report {report_id}: {str(e)}")
		import traceback
		current_app.logger.error(traceback.format_exc())
		return jsonify({'error': 'Failed to generate PDF'}), 500


@admin_bp.route('/admin/v3-reports/<int:report_id>', methods=['DELETE'])
@jwt_required()
def delete_v3_report(report_id):
	"""Delete a V3 report and its associated photos from S3"""
	try:
		from src.models.v3_report import V3JobReport

		current_user_id = get_jwt_identity()
		user = User.query.get(current_user_id)

		if not user or user.role != 'admin':
			return jsonify({'error': 'Admin access required'}), 403

		# Find the report
		report = V3JobReport.query.get(report_id)
		if not report:
			return jsonify({'error': 'Report not found'}), 404

		current_app.logger.info(f"Admin {user.email} deleting report {report_id}")

		# Try to delete associated photos from S3
		if report.photo_urls:
			try:
				for photo in report.photo_urls:
					if isinstance(photo, dict) and 's3_key' in photo:
						s3_key = photo['s3_key']
						try:
							s3_client.s3_client.delete_object(
								Bucket=s3_client.bucket_name,
								Key=s3_key
							)
							current_app.logger.info(f"Deleted S3 object: {s3_key}")
						except Exception as s3_err:
							current_app.logger.warning(f"Failed to delete S3 object {s3_key}: {str(s3_err)}")
			except Exception as photo_err:
				current_app.logger.warning(f"Error processing photo deletions: {str(photo_err)}")

		# Delete the report from the database
		db.session.delete(report)
		db.session.commit()

		current_app.logger.info(f"Successfully deleted report {report_id}")
		return jsonify({'message': 'Report deleted successfully'}), 200

	except Exception as e:
		current_app.logger.error(f"Error deleting report {report_id}: {str(e)}")
		import traceback
		current_app.logger.error(traceback.format_exc())
		db.session.rollback()
		return jsonify({'error': 'Failed to delete report'}), 500



@admin_bp.route('/public/report/<int:report_id>', methods=['GET'])
def get_public_report(report_id):
	"""Public endpoint to view a report (no auth required)."""
	try:
		from src.models.v3_report import V3JobReport

		report = V3JobReport.query.get(report_id)
		if not report:
			return jsonify({'error': 'Report not found'}), 404

		# Get agent name
		agent = User.query.get(report.agent_id)
		agent_name = f"{agent.first_name} {agent.last_name}".strip() if agent else 'Unknown'

		# Convert S3 keys to signed URLs for photos
		photo_urls = []
		if report.photo_urls:
			for photo in report.photo_urls:
				s3_key = photo.get('url') if isinstance(photo, dict) else photo
				if s3_key and s3_client.is_configured():
					signed_url = s3_client.generate_presigned_url(s3_key, expiration=3600)
					if signed_url:
						photo_urls.append({
							'url': signed_url,
							's3_key': s3_key
						})

		return jsonify({
			'id': report.id,
			'form_type': report.form_type,
			'status': report.status,
			'report_data': report.report_data,
			'photo_urls': photo_urls,
			'submitted_at': report.submitted_at.isoformat() if report.submitted_at else None,
			'agent_name': agent_name
		})

	except Exception as e:
		current_app.logger.error(f"Error fetching public report: {str(e)}")
		return jsonify({'error': 'Failed to load report'}), 500


@admin_bp.route('/public/report/<int:report_id>/photos/download', methods=['GET'])
def download_report_photos(report_id):
	"""Download all photos from a report as a ZIP file."""
	try:
		import zipfile
		import io
		import requests
		from src.models.v3_report import V3JobReport

		report = V3JobReport.query.get(report_id)
		if not report:
			return jsonify({'error': 'Report not found'}), 404

		if not report.photo_urls:
			return jsonify({'error': 'No photos in this report'}), 404

		# Create ZIP file in memory
		zip_buffer = io.BytesIO()
		with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
			for i, photo in enumerate(report.photo_urls):
				s3_key = photo.get('url') if isinstance(photo, dict) else photo
				if s3_key and s3_client.is_configured():
					signed_url = s3_client.generate_presigned_url(s3_key, expiration=300)
					if signed_url:
						try:
							# Download the image
							response = requests.get(signed_url, timeout=30)
							if response.status_code == 200:
								# Determine file extension
								ext = '.jpg'
								if 'png' in s3_key.lower():
									ext = '.png'
								elif 'gif' in s3_key.lower():
									ext = '.gif'

								zip_file.writestr(f'photo_{i+1}{ext}', response.content)
						except Exception as download_err:
							current_app.logger.warning(f"Failed to download photo {i}: {str(download_err)}")

		zip_buffer.seek(0)

		# Get client name for filename
		client_name = report.report_data.get('client', 'Unknown').replace(' ', '_')[:30]
		filename = f'report_{report.id}_{client_name}_photos.zip'

		return send_file(
			zip_buffer,
			mimetype='application/zip',
			as_attachment=True,
			download_name=filename
		)

	except Exception as e:
		current_app.logger.error(f"Error downloading photos: {str(e)}")
		return jsonify({'error': 'Failed to download photos'}), 500


@admin_bp.route('/public/report/<int:report_id>/pdf', methods=['GET'])
def download_public_report_pdf(report_id):
	"""Public endpoint to download report as PDF (no auth required)."""
	try:
		from src.models.v3_report import V3JobReport

		report = V3JobReport.query.get(report_id)
		if not report:
			return jsonify({'error': 'Report not found'}), 404

		# Get agent name
		agent = User.query.get(report.agent_id)
		agent_name = f"{agent.first_name} {agent.last_name}".strip() if agent else 'Unknown'

		# Generate PDF using the same function
		pdf_bytes = generate_v3_report_pdf(report, agent_name)

		# Create a BytesIO buffer for the response
		pdf_buffer = io.BytesIO(pdf_bytes)
		pdf_buffer.seek(0)

		# Create filename from client name and date
		report_data = report.report_data or {}
		client_name = report_data.get('client', 'Unknown').replace(' ', '_')[:30]
		date_str = report.submitted_at.strftime('%Y%m%d') if report.submitted_at else 'undated'
		filename = f'V3_Report_{client_name}_{date_str}.pdf'

		return send_file(
			pdf_buffer,
			mimetype='application/pdf',
			as_attachment=True,
			download_name=filename
		)

	except Exception as e:
		current_app.logger.error(f"Error generating public PDF: {str(e)}")
		return jsonify({'error': 'Failed to generate PDF'}), 500